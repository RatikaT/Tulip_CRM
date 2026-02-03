"""
Lead Management Routes
"""
from fastapi import APIRouter, HTTPException, status, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, date, timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.schemas.lead import (
    LeadCreateRequest,
    LeadUpdateRequest,
    LeadResponse,
    LeadListResponse,
    CommentCreateRequest
)
from app.models.lead import Lead, LeadStatus, LeadSource, Trimester, LookingFor, ServiceEnrolled, ServicePartner, ReasonForNoSale
from app.models.audit_log import AuditLog, AuditAction
from app.models.user import User
from app.models.enrollment import Enrollment, ConnectStatus as EnrollmentConnectStatus
from app.utils.enrollment_id import generate_enrollment_id
from app.middleware.auth_middleware import get_current_user, get_current_admin
from app.utils.lead_id import generate_lead_id
from app.database import get_database
import logging
import math
import re

logger = logging.getLogger(__name__)

router = APIRouter()


def lead_to_response(lead: Lead) -> dict:
    """Convert Lead document to response dict"""
    return {
        "id": str(lead.id),
        "lead_id": lead.lead_id,

        # Timestamps
        "created_at": lead.created_at,
        "updated_at": lead.updated_at,

        # Lead Source and Status
        "lead_source": lead.lead_source.value if lead.lead_source else None,
        "lead_creation_date": lead.lead_creation_date,
        "status": lead.status.value if lead.status else None,

        # User Details
        "name": lead.name,
        "email": lead.email,
        "phone_number": lead.phone_number,
        "alternate_mobile_number": lead.alternate_mobile_number,
        "employee_id": lead.employee_id,
        "uhid": lead.uhid,

        # Location Details
        "user_facility": lead.user_facility,
        "city": lead.city,
        "pin_code": lead.pin_code,
        "address": lead.address,

        # Lead Information
        "trimester": lead.trimester.value if lead.trimester else None,
        "looking_for": lead.looking_for.value if lead.looking_for else None,
        "family_member_relation": lead.family_member_relation,
        "package_requested": lead.package_requested,

        # Service Details
        "service_requested": lead.service_requested if lead.service_requested else None,
        "package_name_enrolled": lead.package_name_enrolled,
        "service_partner": lead.service_partner if lead.service_partner else None,
        "provider_location": lead.provider_location,
        "hclhc_spoc": lead.hclhc_spoc,

        # Reason for No Sale
        "reason_for_no_sale": lead.reason_for_no_sale.value if lead.reason_for_no_sale else None,

        # Doctor/Consultation Details
        "doctor_name": lead.doctor_name,
        "doctor_speciality": lead.doctor_speciality,
        "consult_date": lead.consult_date,

        # Medical/Clinical Details
        "visit_id": lead.visit_id,
        "age": lead.age,
        "gender": lead.gender,
        "icd_code": lead.icd_code,
        "diagnosis": lead.diagnosis,
        "investigation_item_name": lead.investigation_item_name,
        "investigation_service_type": lead.investigation_service_type,
        "cug_name": lead.cug_name,

        # Call Tracking
        "number_of_calls": lead.number_of_calls,
        "calls": lead.calls,
        "follow_up_date": lead.follow_up_date,

        # Assignment
        "assigned_to": lead.assigned_to,
        "assigned_to_name": lead.assigned_to_name,
        "assigned_date": lead.assigned_date,
        "reassign_to": lead.reassign_to,
        "reassign_to_name": lead.reassign_to_name,
        "reassigned_date": lead.reassigned_date,

        # Comments
        "comments": lead.comments,

        # System
        "created_by": lead.created_by
    }


@router.get("/stats")
async def get_lead_stats(
    current_user: dict = Depends(get_current_user)
):
    """Get lead statistics for dashboard cards"""
    try:
        logger.info(f"Stats endpoint called by: {current_user.get('full_name')}")

        # MongoDB stores all dates in UTC
        # Server runs in IST (UTC+5:30), so we need to convert IST date boundaries to UTC
        IST_OFFSET = timedelta(hours=5, minutes=30)

        today_local = date.today()  # Local date (IST)
        # IST midnight today = UTC yesterday 18:30
        today_start_utc = datetime.combine(today_local, datetime.min.time()) - IST_OFFSET
        # IST end of today = UTC today 18:29:59
        today_end_utc = datetime.combine(today_local, datetime.max.time()) - IST_OFFSET

        db = get_database()
        is_agent = current_user.get("role") == "agent"
        user_id = current_user.get("user_id", "")

        logger.info(f"Fetching lead stats for user: {current_user.get('full_name')}, role: {current_user.get('role')}")
        logger.info(f"Today (IST): {today_local}, UTC range: {today_start_utc} to {today_end_utc}")

        # Base query
        if is_agent:
            # Agent sees only their assigned leads
            base_query = {
                "is_deleted": False,
                "$or": [
                    {"assigned_to": user_id},
                    {"reassign_to": user_id}
                ]
            }
        else:
            # Admin sees all leads
            base_query = {"is_deleted": False}

        # 1. Total leads
        total = await db.leads.count_documents(base_query)

        # 2. New leads today (created_at is today in IST)
        new_today = await db.leads.count_documents({
            **base_query,
            "created_at": {"$gte": today_start_utc, "$lte": today_end_utc}
        })

        # 3. Follow-ups today (follow_up_date is today in IST)
        follow_up_today = await db.leads.count_documents({
            **base_query,
            "follow_up_date": {"$gte": today_start_utc, "$lte": today_end_utc}
        })

        # 4. Assigned today (for agents - leads assigned or reassigned to them today)
        assigned_today = 0
        if is_agent:
            # Count leads where assigned_date or reassigned_date is today
            assigned_today = await db.leads.count_documents({
                "is_deleted": False,
                "$or": [
                    {"assigned_to": user_id, "assigned_date": {"$gte": today_start_utc, "$lte": today_end_utc}},
                    {"reassign_to": user_id, "reassigned_date": {"$gte": today_start_utc, "$lte": today_end_utc}}
                ]
            })

        logger.info(f"Lead stats - total: {total}, new_today: {new_today}, follow_up_today: {follow_up_today}, assigned_today: {assigned_today}")

        return {
            "total": total,
            "new_today": new_today,
            "follow_up_today": follow_up_today,
            "assigned_today": assigned_today
        }
    except Exception as e:
        logger.error(f"Error in get_lead_stats: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error fetching lead stats: {str(e)}")


@router.get("/bulk-upload/template")
async def get_bulk_upload_template(
    current_user: dict = Depends(get_current_user)
):
    """
    Download a CSV template for bulk lead upload
    """
    # Define template columns
    columns = [
        "name", "phone_number", "email", "uhid", "employee_id",
        "lead_source", "status", "trimester", "looking_for",
        "city", "address", "pin_code", "service_partner", "provider_location",
        "service_requested", "package_name_enrolled", "hclhc_spoc",
        "doctor_name", "doctor_speciality", "follow_up_date", "alternate_mobile_number",
        "assigned_to"
    ]

    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    # Add a sample row
    writer.writerow([
        "John Doe", "9876543210", "john@example.com", "UHID001", "EMP001",
        "Website", "Enquiry Lead", "Trimester 1", "Self",
        "Delhi", "123 Main St", "110001", "Apollo Cradle", "Kondapur",
        "Tulip Antenatal", "Basic Package", "Agent Name",
        "Dr. Smith", "Gynecology", "2026-02-15", "9876543211",
        "Richa"
    ])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=leads_bulk_upload_template.csv"}
    )


@router.post("/bulk-upload")
async def bulk_upload_leads(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_admin)
):
    """
    Bulk upload leads from CSV file (Admin only)
    """
    if not file.filename.endswith('.csv'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV files are supported"
        )

    db = get_database()
    errors = []
    created_count = 0
    total_rows = 0

    try:
        # Read file content
        contents = await file.read()
        decoded = contents.decode('utf-8-sig')  # Handle BOM
        reader = csv.DictReader(io.StringIO(decoded))

        for row_num, row in enumerate(reader, start=2):  # Start at 2 (row 1 is header)
            total_rows += 1
            try:
                # Get identifier fields
                name = row.get('name', '').strip()
                phone = row.get('phone_number', '').strip()
                email = row.get('email', '').strip()
                uhid = row.get('uhid', '').strip()

                # Required: At least one of UHID, Contact No., or Email
                if not uhid and not phone and not email:
                    errors.append({"row": row_num, "error": "At least one of UHID, Contact No., or Email is required"})
                    continue

                # Validate phone format if provided
                if phone and (len(phone) != 10 or not phone.isdigit()):
                    errors.append({"row": row_num, "error": f"Invalid phone number format: {phone} (must be 10 digits)"})
                    continue

                # Parse lead source
                lead_source_str = row.get('lead_source', '').strip()
                lead_source = None
                if lead_source_str:
                    source_map = {
                        'prescription dump': LeadSource.PRESCRIPTION_DUMP,
                        'in clinic-gynae consult': LeadSource.IN_CLINIC_GYNAE_CONSULT,
                        'in clinic-other consults': LeadSource.IN_CLINIC_OTHER_CONSULTS,
                        'in clinic-walk in': LeadSource.IN_CLINIC_WALK_IN,
                        'ama': LeadSource.AMA,
                        'bewell': LeadSource.BEWELL,
                        'events': LeadSource.EVENTS,
                        'call': LeadSource.CALL,
                        'others': LeadSource.OTHERS,
                        'bump day': LeadSource.BUMP_DAY,
                        'whatsapp': LeadSource.WHATSAPP,
                        'mail': LeadSource.MAIL,
                        'tele-consultation': LeadSource.TELE_CONSULTATION,
                        'website': LeadSource.WEBSITE,
                        'habit banner': LeadSource.HABIT_BANNER,
                        # Backward compatibility mappings
                        'wa': LeadSource.WHATSAPP,
                        'sms': LeadSource.OTHERS,
                        'emr': LeadSource.OTHERS,
                        'other': LeadSource.OTHERS
                    }
                    lead_source = source_map.get(lead_source_str.lower(), LeadSource.OTHERS)
                else:
                    lead_source = LeadSource.OTHERS

                # Parse status
                status_str = row.get('status', '').strip()
                lead_status = LeadStatus.ENQUIRY_LEAD
                if status_str:
                    status_map = {
                        'not interested': LeadStatus.NOT_INTERESTED,
                        'enquiry lead': LeadStatus.ENQUIRY_LEAD,
                        'lead closed-no response': LeadStatus.LEAD_CLOSED_NO_RESPONSE,
                        'lead closed - no response': LeadStatus.LEAD_CLOSED_NO_RESPONSE,
                        'enrolled': LeadStatus.ENROLLED,
                        'follow up-in process': LeadStatus.FOLLOWUP_IN_PROCESS,
                        'follow up-no response': LeadStatus.FOLLOWUP_NO_RESPONSE,
                        'duplicate': LeadStatus.DUPLICATE,
                        # Backward compatibility
                        'new': LeadStatus.ENQUIRY_LEAD,
                        'interested': LeadStatus.ENQUIRY_LEAD,
                        'followup required': LeadStatus.FOLLOWUP_IN_PROCESS,
                        'no response': LeadStatus.FOLLOWUP_NO_RESPONSE,
                    }
                    lead_status = status_map.get(status_str.lower(), LeadStatus.ENQUIRY_LEAD)

                # Parse trimester
                trimester_str = row.get('trimester', '').strip()
                trimester = None
                if trimester_str:
                    trimester_map = {
                        'trimester 1': Trimester.TRIMESTER_1,
                        'trimester1': Trimester.TRIMESTER_1,
                        '1st': Trimester.TRIMESTER_1,
                        '1': Trimester.TRIMESTER_1,
                        'trimester 2': Trimester.TRIMESTER_2,
                        'trimester2': Trimester.TRIMESTER_2,
                        '2nd': Trimester.TRIMESTER_2,
                        '2': Trimester.TRIMESTER_2,
                        'trimester 3': Trimester.TRIMESTER_3,
                        'trimester3': Trimester.TRIMESTER_3,
                        '3rd': Trimester.TRIMESTER_3,
                        '3': Trimester.TRIMESTER_3,
                        'not conceived': Trimester.NOT_CONCEIVED,
                        'notconceived': Trimester.NOT_CONCEIVED,
                    }
                    trimester = trimester_map.get(trimester_str.lower())

                # Parse looking_for
                looking_for_str = row.get('looking_for', '').strip()
                looking_for = None
                if looking_for_str:
                    if looking_for_str.lower() == 'self':
                        looking_for = LookingFor.SELF
                    elif 'family' in looking_for_str.lower():
                        looking_for = LookingFor.FAMILY_MEMBER

                # Parse service_requested
                service_str = row.get('service_requested', '').strip()
                service_requested = None
                if service_str:
                    service_map = {
                        'preconception': ServiceEnrolled.PRE_CONCEPTION,
                        'pre conception': ServiceEnrolled.PRE_CONCEPTION,
                        'antenatal': ServiceEnrolled.ANTENATAL,
                        'maternitywellness': ServiceEnrolled.MATERNITY_WELLNESS,
                        'maternity wellness': ServiceEnrolled.MATERNITY_WELLNESS,
                    }
                    service_requested = service_map.get(service_str.lower())

                # Parse service_partner
                partner_str = row.get('service_partner', '').strip()
                service_partner = None
                if partner_str:
                    partner_map = {
                        'motherhood': ServicePartner.MOTHERHOOD,
                        'rainbow': ServicePartner.RAINBOW,
                        'fortis': ServicePartner.FORTIS,
                        'apollo cradle': ServicePartner.APOLLO_CRADLE,
                        'cloud 9': ServicePartner.CLOUD_9,
                        'hcl healthcare': ServicePartner.HCL_HEALTHCARE,
                        'mamily': ServicePartner.MAMILY,
                        'others': ServicePartner.OTHERS,
                    }
                    service_partner = partner_map.get(partner_str.lower(), ServicePartner.OTHERS)

                # Parse reason_for_no_sale
                reason_str = row.get('reason_for_no_sale', '').strip()
                reason_for_no_sale = None
                if reason_str:
                    reason_map = {
                        'already taking service outside': ReasonForNoSale.ALREADY_TAKING_SERVICE_OUTSIDE,
                        'location not suitable': ReasonForNoSale.LOCATION_NOT_SUITABLE,
                        'different service provider required-brand': ReasonForNoSale.DIFFERENT_SERVICE_PROVIDER_REQUIRED,
                        'travelling to native place for delivery': ReasonForNoSale.TRAVELLING_TO_NATIVE,
                        'package cost': ReasonForNoSale.PACKAGE_COST,
                        'only delivery package required': ReasonForNoSale.ONLY_DELIVERY_PACKAGE,
                        'package inadequate': ReasonForNoSale.PACKAGE_INADEQUATE,
                        'miscarriage': ReasonForNoSale.MISCARRIAGE,
                        'looking for other hclh services': ReasonForNoSale.LOOKING_FOR_OTHER_HCLH,
                        'others': ReasonForNoSale.OTHERS,
                    }
                    reason_for_no_sale = reason_map.get(reason_str.lower())

                # Parse dates
                lead_creation_date = None
                lcd_str = row.get('lead_creation_date', '').strip()
                if lcd_str:
                    try:
                        lead_creation_date = date.fromisoformat(lcd_str)
                    except:
                        pass

                consult_date_val = None
                cd_str = row.get('consult_date', '').strip()
                if cd_str:
                    try:
                        consult_date_val = date.fromisoformat(cd_str)
                    except:
                        pass

                follow_up_date = None
                fud_str = row.get('follow_up_date', '').strip()
                if fud_str:
                    try:
                        follow_up_date = datetime.fromisoformat(fud_str)
                    except:
                        pass

                # Parse age
                age_val = None
                age_str = row.get('age', '').strip()
                if age_str:
                    try:
                        age_val = int(age_str)
                    except:
                        pass

                # Parse assigned_to - look up user by name
                assigned_to_str = row.get('assigned_to', '').strip()
                assigned_to_id = current_user["user_id"]
                assigned_to_name = current_user["full_name"]
                if assigned_to_str:
                    # Try to find user by full_name (case-insensitive)
                    assigned_user = await User.find_one({"full_name": {"$regex": f"^{re.escape(assigned_to_str)}$", "$options": "i"}})
                    if assigned_user:
                        assigned_to_id = str(assigned_user.id)
                        assigned_to_name = assigned_user.full_name

                # Generate lead ID
                lead_id = await generate_lead_id(db)

                # Create lead
                lead = Lead(
                    lead_id=lead_id,
                    lead_source=lead_source,
                    lead_creation_date=lead_creation_date,
                    status=lead_status,
                    name=name or "Unknown",
                    email=email or None,
                    phone_number=phone or None,
                    employee_id=row.get('employee_id', '').strip() or None,
                    uhid=uhid or None,
                    user_facility=row.get('user_facility', '').strip() or None,
                    city=row.get('city', '').strip() or None,
                    pin_code=row.get('pin_code', '').strip() or None,
                    address=row.get('address', '').strip() or None,
                    trimester=trimester,
                    looking_for=looking_for,
                    package_requested=row.get('package_requested', '').strip() or None,
                    service_requested=service_requested,
                    package_name_enrolled=row.get('package_name_enrolled', '').strip() or None,
                    service_partner=service_partner,
                    provider_location=row.get('provider_location', '').strip() or None,
                    hclhc_spoc=row.get('hclhc_spoc', '').strip() or None,
                    reason_for_no_sale=reason_for_no_sale,
                    doctor_name=row.get('doctor_name', '').strip() or None,
                    doctor_speciality=row.get('doctor_speciality', '').strip() or None,
                    consult_date=consult_date_val,
                    follow_up_date=follow_up_date,
                    # Medical/Clinical Details
                    visit_id=row.get('visit_id', '').strip() or None,
                    age=age_val,
                    gender=row.get('gender', '').strip() or None,
                    icd_code=row.get('icd_code', '').strip() or None,
                    diagnosis=row.get('diagnosis', '').strip() or None,
                    investigation_item_name=row.get('investigation_item_name', '').strip() or None,
                    investigation_service_type=row.get('investigation_service_type', '').strip() or None,
                    cug_name=row.get('cug_name', '').strip() or None,
                    created_by=current_user["user_id"],
                    assigned_to=assigned_to_id,
                    assigned_to_name=assigned_to_name,
                    number_of_calls=1,
                    calls=[],
                    comments=[],
                    is_deleted=False
                )

                await lead.insert()
                created_count += 1

            except Exception as e:
                errors.append({"row": row_num, "error": str(e)})

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to process file: {str(e)}"
        )

    logger.info(f"Bulk upload by {current_user['email']}: {created_count} created, {len(errors)} errors")

    return {
        "success": len(errors) == 0 or created_count > 0,
        "message": f"Processed {total_rows} rows. Created {created_count} leads." +
                   (f" {len(errors)} errors." if errors else ""),
        "total_rows": total_rows,
        "created": created_count,
        "errors": errors[:20]  # Limit errors returned
    }


@router.get("", response_model=LeadListResponse)
async def get_leads(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=100),
    status: Optional[List[str]] = Query(None),
    lead_source: Optional[List[str]] = Query(None),
    uhid: Optional[List[str]] = Query(None),
    package_requested: Optional[List[str]] = Query(None),
    city: Optional[str] = None,
    assigned_to: Optional[str] = None,
    reassign_to: Optional[str] = None,
    created_date_from: Optional[str] = None,
    created_date_to: Optional[str] = None,
    next_follow_up_date: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Get paginated list of leads with filters
    Agents can only see leads assigned to them or reassigned to them
    Supports multi-select filters for status, lead_source, uhid, package_requested
    """
    # Build query
    query = {"is_deleted": False}

    # Agent restriction: only see leads assigned or reassigned to them
    if current_user["role"] == "agent":
        user_id = current_user["user_id"]
        query["$or"] = [
            {"assigned_to": user_id},
            {"reassign_to": user_id}
        ]

    # Multi-select filters (support arrays with $in)
    if status and len(status) > 0:
        query["status"] = {"$in": status}
    if lead_source and len(lead_source) > 0:
        query["lead_source"] = {"$in": lead_source}
    if uhid and len(uhid) > 0:
        query["uhid"] = {"$in": uhid}
    if package_requested and len(package_requested) > 0:
        query["package_requested"] = {"$in": package_requested}

    # Single value filters
    if city:
        query["city"] = {"$regex": re.escape(city), "$options": "i"}

    # Assigned To and Reassign To filters with OR logic
    if assigned_to and reassign_to:
        # If both filters are set, use OR logic
        assignment_conditions = [
            {"assigned_to": assigned_to},
            {"reassign_to": reassign_to}
        ]
        # If there's already an $or in query (from agent restriction), we need to handle it
        if "$or" in query:
            # Combine with existing $or using $and
            existing_or = query.pop("$or")
            if "$and" not in query:
                query["$and"] = []
            query["$and"].append({"$or": existing_or})
            query["$and"].append({"$or": assignment_conditions})
        else:
            query["$or"] = assignment_conditions
    elif assigned_to:
        query["assigned_to"] = assigned_to
    elif reassign_to:
        query["reassign_to"] = reassign_to

    # Date range filter for created_at
    if created_date_from or created_date_to:
        created_at_filter = {}
        if created_date_from:
            try:
                from_date = datetime.strptime(created_date_from, "%Y-%m-%d").date()
                created_at_filter["$gte"] = datetime.combine(from_date, datetime.min.time())
            except ValueError:
                pass
        if created_date_to:
            try:
                to_date = datetime.strptime(created_date_to, "%Y-%m-%d").date()
                created_at_filter["$lte"] = datetime.combine(to_date, datetime.max.time())
            except ValueError:
                pass
        if created_at_filter:
            query["created_at"] = created_at_filter

    # Next follow up date filter (single day)
    if next_follow_up_date:
        try:
            filter_date = datetime.strptime(next_follow_up_date, "%Y-%m-%d").date()
            day_start = datetime.combine(filter_date, datetime.min.time())
            day_end = datetime.combine(filter_date, datetime.max.time())
            query["follow_up_date"] = {"$gte": day_start, "$lte": day_end}
        except ValueError:
            pass

    # Search by name, phone, or lead_id (escape regex special chars for security)
    if search:
        escaped_search = re.escape(search)
        # For agents, we need to combine search with their assignment filter
        search_conditions = [
            {"name": {"$regex": escaped_search, "$options": "i"}},
            {"phone_number": {"$regex": escaped_search}},
            {"lead_id": {"$regex": escaped_search, "$options": "i"}},
            {"email": {"$regex": escaped_search, "$options": "i"}}
        ]
        if current_user["role"] == "agent":
            # Combine agent's assignment filter with search
            user_id = current_user["user_id"]
            base_filters = [
                {"is_deleted": False},
                {"$or": [{"assigned_to": user_id}, {"reassign_to": user_id}]},
                {"$or": search_conditions}
            ]
            # Add other filters if present
            if status and len(status) > 0:
                base_filters.append({"status": {"$in": status}})
            if lead_source and len(lead_source) > 0:
                base_filters.append({"lead_source": {"$in": lead_source}})
            if uhid and len(uhid) > 0:
                base_filters.append({"uhid": {"$in": uhid}})
            if package_requested and len(package_requested) > 0:
                base_filters.append({"package_requested": {"$in": package_requested}})
            if city:
                base_filters.append({"city": {"$regex": re.escape(city), "$options": "i"}})
            # Handle assigned_to and reassign_to with OR logic
            if assigned_to and reassign_to:
                base_filters.append({"$or": [{"assigned_to": assigned_to}, {"reassign_to": reassign_to}]})
            elif assigned_to:
                base_filters.append({"assigned_to": assigned_to})
            elif reassign_to:
                base_filters.append({"reassign_to": reassign_to})
            if "created_at" in query:
                base_filters.append({"created_at": query["created_at"]})
            if "follow_up_date" in query:
                base_filters.append({"follow_up_date": query["follow_up_date"]})
            query = {"$and": base_filters}
        else:
            query["$or"] = search_conditions

    # Count total
    total = await Lead.find(query).count()
    pages = math.ceil(total / per_page) if total > 0 else 1

    # Get paginated results
    skip = (page - 1) * per_page
    leads = await Lead.find(query).sort("-created_at").skip(skip).limit(per_page).to_list()

    return {
        "leads": [lead_to_response(lead) for lead in leads],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": pages
    }


@router.get("/export/excel")
async def export_leads_excel(
    current_user: dict = Depends(get_current_admin)
):
    """
    Export all leads with audit trail to Excel (Admin only)
    """
    # Create workbook
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== Sheet 1: Leads =====
    ws_leads = wb.active
    ws_leads.title = "Leads"

    # Lead headers
    lead_headers = [
        "Lead ID", "Name", "Email", "Phone Number", "Employee ID", "UHID",
        "Status", "Lead Source", "Lead Creation Date", "Trimester", "Looking For",
        "User Facility", "City", "Pin Code", "Address",
        "Package Requested", "Service Enrolled", "Package Name Enrolled",
        "Service (Partner)", "Provider Location", "HCLHC SPOC", "Reason for No Sale",
        "Doctor Name", "Doctor Speciality", "Consult Date",
        "Visit ID", "Age", "Gender", "ICD Code", "Diagnosis",
        "Investigation Item Name", "Investigation Service Type", "CUG Name",
        "Number of Calls", "Follow Up Date",
        "Assigned To", "Reassign To",
        "Created At", "Updated At", "Created By"
    ]

    # Write headers
    for col, header in enumerate(lead_headers, 1):
        cell = ws_leads.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Fetch all leads
    leads = await Lead.find(Lead.is_deleted == False).sort("-created_at").to_list()

    # Write lead data
    for row_num, lead in enumerate(leads, 2):
        row_data = [
            lead.lead_id,
            lead.name,
            lead.email,
            lead.phone_number,
            lead.employee_id,
            lead.uhid,
            lead.status.value if lead.status else None,
            lead.lead_source.value if lead.lead_source else None,
            str(lead.lead_creation_date) if lead.lead_creation_date else None,
            lead.trimester.value if lead.trimester else None,
            lead.looking_for.value if lead.looking_for else None,
            lead.user_facility,
            lead.city,
            lead.pin_code,
            lead.address,
            lead.package_requested,
            lead.service_requested if lead.service_requested else None,
            lead.package_name_enrolled,
            lead.service_partner if lead.service_partner else None,
            lead.provider_location,
            lead.hclhc_spoc,
            lead.reason_for_no_sale.value if lead.reason_for_no_sale else None,
            lead.doctor_name,
            lead.doctor_speciality,
            str(lead.consult_date) if lead.consult_date else None,
            lead.visit_id,
            lead.age,
            lead.gender,
            lead.icd_code,
            lead.diagnosis,
            lead.investigation_item_name,
            lead.investigation_service_type,
            lead.cug_name,
            lead.number_of_calls,
            lead.follow_up_date.strftime("%Y-%m-%d %H:%M") if lead.follow_up_date else None,
            lead.assigned_to_name,
            getattr(lead, 'reassign_to_name', None),
            lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else None,
            lead.updated_at.strftime("%Y-%m-%d %H:%M") if lead.updated_at else None,
            lead.created_by
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws_leads.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column widths for leads sheet
    for col in ws_leads.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_leads.column_dimensions[column].width = adjusted_width

    # ===== Sheet 2: Calls =====
    ws_calls = wb.create_sheet("Calls")

    call_headers = ["Lead ID", "Name", "Call Number", "Date & Time", "Summary"]
    for col, header in enumerate(call_headers, 1):
        cell = ws_calls.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    call_row = 2
    for lead in leads:
        if lead.calls:
            for call in lead.calls:
                ws_calls.cell(row=call_row, column=1, value=lead.lead_id).border = thin_border
                ws_calls.cell(row=call_row, column=2, value=lead.name).border = thin_border
                ws_calls.cell(row=call_row, column=3, value=call.get('call_number', '')).border = thin_border
                ws_calls.cell(row=call_row, column=4, value=call.get('date_time', '')).border = thin_border
                ws_calls.cell(row=call_row, column=5, value=call.get('summary', '')).border = thin_border
                call_row += 1

    # Auto-adjust column widths for calls sheet
    for col in ws_calls.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws_calls.column_dimensions[column].width = adjusted_width

    # ===== Sheet 3: Comments =====
    ws_comments = wb.create_sheet("Comments")

    comment_headers = ["Lead ID", "Name", "Comment", "Created By", "Created At"]
    for col, header in enumerate(comment_headers, 1):
        cell = ws_comments.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    comment_row = 2
    for lead in leads:
        if lead.comments:
            for comment in lead.comments:
                ws_comments.cell(row=comment_row, column=1, value=lead.lead_id).border = thin_border
                ws_comments.cell(row=comment_row, column=2, value=lead.name).border = thin_border
                ws_comments.cell(row=comment_row, column=3, value=comment.get('text', '')).border = thin_border
                ws_comments.cell(row=comment_row, column=4, value=comment.get('created_by_name', '')).border = thin_border
                created_at = comment.get('created_at')
                if created_at:
                    if isinstance(created_at, datetime):
                        created_at = created_at.strftime("%Y-%m-%d %H:%M")
                ws_comments.cell(row=comment_row, column=5, value=created_at).border = thin_border
                comment_row += 1

    # Auto-adjust column widths for comments sheet
    for col in ws_comments.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws_comments.column_dimensions[column].width = adjusted_width

    # ===== Sheet 4: Audit Trail =====
    ws_audit = wb.create_sheet("Audit Trail")

    audit_headers = ["Lead ID", "User", "Email", "Action", "Field", "Old Value", "New Value", "Timestamp"]
    for col, header in enumerate(audit_headers, 1):
        cell = ws_audit.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Fetch all audit logs
    audit_logs = await AuditLog.find().sort("-timestamp").to_list()

    audit_row = 2
    for log in audit_logs:
        if log.changes:
            for change in log.changes:
                ws_audit.cell(row=audit_row, column=1, value=log.lead_id).border = thin_border
                ws_audit.cell(row=audit_row, column=2, value=log.user_name).border = thin_border
                ws_audit.cell(row=audit_row, column=3, value=log.user_email).border = thin_border
                ws_audit.cell(row=audit_row, column=4, value=log.action.value if hasattr(log.action, 'value') else str(log.action)).border = thin_border
                ws_audit.cell(row=audit_row, column=5, value=change.get('field', '')).border = thin_border
                ws_audit.cell(row=audit_row, column=6, value=str(change.get('old_value', ''))).border = thin_border
                ws_audit.cell(row=audit_row, column=7, value=str(change.get('new_value', ''))).border = thin_border
                ws_audit.cell(row=audit_row, column=8, value=log.timestamp.strftime("%Y-%m-%d %H:%M") if log.timestamp else None).border = thin_border
                audit_row += 1
        else:
            ws_audit.cell(row=audit_row, column=1, value=log.lead_id).border = thin_border
            ws_audit.cell(row=audit_row, column=2, value=log.user_name).border = thin_border
            ws_audit.cell(row=audit_row, column=3, value=log.user_email).border = thin_border
            ws_audit.cell(row=audit_row, column=4, value=log.action.value if hasattr(log.action, 'value') else str(log.action)).border = thin_border
            ws_audit.cell(row=audit_row, column=8, value=log.timestamp.strftime("%Y-%m-%d %H:%M") if log.timestamp else None).border = thin_border
            audit_row += 1

    # Auto-adjust column widths for audit sheet
    for col in ws_audit.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_audit.column_dimensions[column].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with timestamp
    filename = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    logger.info(f"Excel export generated by {current_user['email']}: {len(leads)} leads, {len(audit_logs)} audit entries")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{lead_id}", response_model=LeadResponse)
async def get_lead(
    lead_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Get single lead by ID
    Agents can only view leads assigned to them or reassigned to them
    """
    lead = await Lead.find_one(Lead.lead_id == lead_id, Lead.is_deleted == False)

    if not lead:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lead not found"
        )

    # Agent restriction: only view leads assigned or reassigned to them
    if current_user["role"] == "agent":
        user_id = current_user["user_id"]
        if lead.assigned_to != user_id and lead.reassign_to != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have permission to view this lead"
            )

    return lead_to_response(lead)


@router.post("", response_model=LeadResponse)
async def create_lead(
    request: LeadCreateRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Create a new lead (All authenticated users)
    All users can specify assignment when creating a lead
    """
    logger.debug(f"Creating lead: name={request.name}, phone={request.phone_number}, by user={current_user['email']}")
    db = get_database()

    # Generate unique LeadID
    lead_id = await generate_lead_id(db)

    # Get assigned user name if assigned, otherwise default to current user
    assigned_to = request.assigned_to if request.assigned_to else current_user["user_id"]
    assigned_to_name = None
    if request.assigned_to:
        assigned_user = await User.get(request.assigned_to)
        if assigned_user:
            assigned_to_name = assigned_user.full_name
    else:
        assigned_to_name = current_user["full_name"]

    # Create lead with new fields
    lead = Lead(
        lead_id=lead_id,
        lead_source=request.lead_source,
        lead_creation_date=request.lead_creation_date,
        name=request.name,
        email=request.email,
        phone_number=request.phone_number,
        alternate_mobile_number=request.alternate_mobile_number,
        employee_id=request.employee_id,
        uhid=request.uhid,
        user_facility=request.user_facility,
        city=request.city,
        pin_code=request.pin_code,
        address=request.address,
        trimester=request.trimester,
        looking_for=request.looking_for,
        family_member_relation=request.family_member_relation,
        package_requested=request.package_requested,
        service_requested=request.service_requested,
        package_name_enrolled=request.package_name_enrolled,
        service_partner=request.service_partner,
        provider_location=request.provider_location,
        hclhc_spoc=request.hclhc_spoc,
        reason_for_no_sale=request.reason_for_no_sale,
        doctor_speciality=request.doctor_speciality,
        follow_up_date=request.follow_up_date,
        # Medical/Clinical Details
        visit_id=request.visit_id,
        age=request.age,
        gender=request.gender,
        icd_code=request.icd_code,
        diagnosis=request.diagnosis,
        investigation_item_name=request.investigation_item_name,
        investigation_service_type=request.investigation_service_type,
        cug_name=request.cug_name,
        assigned_to=assigned_to,
        assigned_to_name=assigned_to_name,
        assigned_date=datetime.utcnow(),  # Set assigned date on creation
        created_by=current_user["user_id"],
        number_of_calls=1,
        calls=[],
        comments=[]
    )

    await lead.insert()

    # Create audit log
    audit = AuditLog(
        lead_id=lead_id,
        user_id=current_user["user_id"],
        user_email=current_user["email"],
        user_name=current_user["full_name"],
        action=AuditAction.CREATED,
        changes=[{"field": "lead", "old_value": None, "new_value": "created"}]
    )
    await audit.insert()

    logger.info(f"Lead created: {lead_id} by {current_user['email']}")

    return lead_to_response(lead)


@router.put("/{lead_id}", response_model=LeadResponse)
async def update_lead(
    lead_id: str,
    request: LeadUpdateRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Update a lead
    Agents can only update leads assigned/reassigned to them with limited fields
    Admins can update all fields on all leads
    """
    logger.debug(f"Updating lead {lead_id} by user={current_user['email']}, role={current_user['role']}")
    lead = await Lead.find_one(Lead.lead_id == lead_id, Lead.is_deleted == False)

    if not lead:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lead not found"
        )

    # Agent restriction: only update leads assigned or reassigned to them
    if current_user["role"] == "agent":
        user_id = current_user["user_id"]
        if lead.assigned_to != user_id and lead.reassign_to != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have permission to update this lead"
            )

    # Track changes for audit
    changes = []
    update_data = request.model_dump(exclude_unset=True)

    # Agents cannot update assignment fields
    if current_user["role"] == "agent":
        restricted_fields = ["assigned_to", "assigned_to_name"]
        update_data = {k: v for k, v in update_data.items() if k not in restricted_fields}

    # Look up reassign_to_name when reassign_to is provided
    if "reassign_to" in update_data and update_data["reassign_to"]:
        reassign_user = await User.get(update_data["reassign_to"])
        if reassign_user:
            update_data["reassign_to_name"] = reassign_user.full_name
        else:
            update_data["reassign_to_name"] = None
        # Set reassigned_date when reassign_to changes
        update_data["reassigned_date"] = datetime.utcnow()

    # Look up assigned_to_name when assigned_to is provided (for admins)
    if "assigned_to" in update_data and update_data["assigned_to"]:
        assigned_user = await User.get(update_data["assigned_to"])
        if assigned_user:
            update_data["assigned_to_name"] = assigned_user.full_name
        # Set assigned_date when assigned_to changes
        update_data["assigned_date"] = datetime.utcnow()

    # Helper function to format calls for audit
    def format_calls_for_audit(calls_list):
        if not calls_list:
            return "No calls"
        formatted = []
        for call in calls_list:
            if isinstance(call, dict):
                call_num = call.get('call_number', '?')
                date_time = call.get('date_time', 'No date')
                summary = call.get('summary', '') or 'No summary'
                # Truncate summary if too long
                if len(summary) > 50:
                    summary = summary[:50] + '...'
                formatted.append(f"Call {call_num}: {date_time[:16] if len(str(date_time)) > 16 else date_time} - '{summary}'")
            else:
                # Handle CallEntry objects
                call_num = getattr(call, 'call_number', '?')
                date_time = getattr(call, 'date_time', 'No date')
                summary = getattr(call, 'summary', '') or 'No summary'
                if len(summary) > 50:
                    summary = summary[:50] + '...'
                formatted.append(f"Call {call_num}: {str(date_time)[:16]} - '{summary}'")
        return "; ".join(formatted)

    # Helper to check if a date is in the past
    def is_past_date(date_value):
        if not date_value:
            return False
        try:
            if isinstance(date_value, str):
                # Parse ISO format datetime string
                from dateutil import parser
                dt = parser.parse(date_value)
            elif isinstance(date_value, datetime):
                dt = date_value
            else:
                return False
            return dt.date() < datetime.utcnow().date()
        except:
            return False

    # Helper to compare calls and generate detailed changes
    def get_call_changes(old_calls, new_calls):
        call_changes = []
        old_calls = old_calls or []
        new_calls = new_calls or []
        past_edit_errors = []

        # Convert to list of dicts for comparison
        old_dict = {}
        for c in old_calls:
            if isinstance(c, dict):
                old_dict[c.get('call_number')] = c
            else:
                old_dict[getattr(c, 'call_number', None)] = {
                    'call_number': getattr(c, 'call_number', None),
                    'date_time': str(getattr(c, 'date_time', '')),
                    'summary': getattr(c, 'summary', '')
                }

        new_dict = {}
        for c in new_calls:
            if isinstance(c, dict):
                new_dict[c.get('call_number')] = c
            else:
                new_dict[getattr(c, 'call_number', None)] = {
                    'call_number': getattr(c, 'call_number', None),
                    'date_time': str(getattr(c, 'date_time', '')),
                    'summary': getattr(c, 'summary', '')
                }

        # Find new calls
        for call_num, new_call in new_dict.items():
            if call_num not in old_dict:
                summary = new_call.get('summary', '') or 'No summary'
                if len(summary) > 50:
                    summary = summary[:50] + '...'
                call_changes.append({
                    "field": f"Call {call_num}",
                    "old_value": None,
                    "new_value": f"Added - Date: {str(new_call.get('date_time', ''))[:16]}, Summary: '{summary}'"
                })
            else:
                old_call = old_dict[call_num]
                old_date_time = old_call.get('date_time', '')

                # Check if trying to edit summary of a past date call
                old_summary = old_call.get('summary', '') or ''
                new_summary = new_call.get('summary', '') or ''
                if old_summary != new_summary and is_past_date(old_date_time):
                    past_edit_errors.append(f"Call {call_num}")
                    continue

                # Check if summary changed
                new_summary = new_call.get('summary', '') or ''
                if old_summary != new_summary:
                    old_display = old_summary[:50] + '...' if len(old_summary) > 50 else old_summary or '(empty)'
                    new_display = new_summary[:50] + '...' if len(new_summary) > 50 else new_summary or '(empty)'
                    call_changes.append({
                        "field": f"Call {call_num} Summary",
                        "old_value": old_display,
                        "new_value": new_display
                    })
                # Check if date changed
                old_date = str(old_call.get('date_time', ''))[:16]
                new_date = str(new_call.get('date_time', ''))[:16]
                if old_date != new_date:
                    call_changes.append({
                        "field": f"Call {call_num} Date",
                        "old_value": old_date,
                        "new_value": new_date
                    })

        # Find deleted calls
        for call_num in old_dict:
            if call_num not in new_dict:
                call_changes.append({
                    "field": f"Call {call_num}",
                    "old_value": "Existed",
                    "new_value": "Deleted"
                })

        return call_changes, past_edit_errors

    # Apply updates
    for field, new_value in update_data.items():
        old_value = getattr(lead, field, None)

        # Special handling for calls field
        if field == "calls":
            call_changes, past_errors = get_call_changes(old_value, new_value)
            if past_errors:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Cannot edit summaries of past date calls: {', '.join(past_errors)}"
                )
            changes.extend(call_changes)
            setattr(lead, field, new_value)
            continue

        # Skip number_of_calls since we track individual call changes
        if field == "number_of_calls":
            setattr(lead, field, new_value)
            continue

        # Handle enum values
        if hasattr(old_value, 'value'):
            old_value = old_value.value

        # Convert enum strings to enum values for certain fields
        if field in ["lead_source", "status", "trimester", "looking_for", "service_requested", "service_partner", "reason_for_no_sale"] and new_value:
            if isinstance(new_value, str):
                try:
                    if field == "lead_source":
                        new_value = LeadSource(new_value)
                    elif field == "status":
                        new_value = LeadStatus(new_value)
                    elif field == "trimester":
                        new_value = Trimester(new_value)
                    elif field == "looking_for":
                        new_value = LookingFor(new_value)
                    elif field == "service_requested":
                        new_value = ServiceEnrolled(new_value)
                    elif field == "service_partner":
                        new_value = ServicePartner(new_value)
                    elif field == "reason_for_no_sale":
                        new_value = ReasonForNoSale(new_value)
                except ValueError:
                    pass  # Keep as string if conversion fails

        if old_value != new_value:
            changes.append({
                "field": field,
                "old_value": old_value,
                "new_value": new_value.value if hasattr(new_value, 'value') else new_value
            })
            setattr(lead, field, new_value)

    # Update metadata
    lead.updated_at = datetime.utcnow()
    lead.last_modified_by = current_user["user_id"]

    await lead.save()

    # Auto-enrollment: Create enrollment when status changes to "Enrolled"
    status_change = next((c for c in changes if c["field"] == "status" and c["new_value"] == LeadStatus.ENROLLED.value), None)
    if status_change:
        try:
            db = get_database()
            enrollment_id = await generate_enrollment_id(db)

            # Map lead's service_partner to enrollment's service_partner
            service_partner_value = lead.service_partner if lead.service_partner else None

            # Map trimester
            trimester_value = None
            if lead.trimester:
                from app.models.enrollment import Trimester as EnrollmentTrimester
                try:
                    trimester_value = EnrollmentTrimester(lead.trimester.value)
                except ValueError:
                    pass

            # Log lead values being mapped to enrollment
            logger.info(f"Creating enrollment from lead {lead_id} with values: uhid={lead.uhid}, service_requested={lead.service_requested}, package_requested={lead.package_requested}, provider_location={lead.provider_location}")

            enrollment = Enrollment(
                enrollment_id=enrollment_id,
                linked_lead_id=lead.lead_id,
                subscriber_name=lead.name,
                employee_id=lead.employee_id or "",
                phone_number=lead.phone_number,
                email=lead.email,
                uhid=lead.uhid,
                trimester=trimester_value,
                doctor_name=lead.doctor_name,
                service_enrolled=lead.service_requested,
                package_name_enrolled=lead.package_requested,
                service_partner=service_partner_value,
                partner_centre_selected=lead.provider_location,
                hclhc_spoc=lead.hclhc_spoc,
                connect_status=EnrollmentConnectStatus.CONNECTED,
                created_by=current_user["user_id"],
                created_by_name=current_user["full_name"],
                assigned_to=current_user["user_id"],
                assigned_to_name=current_user["full_name"],
            )
            await enrollment.insert()
            logger.info(f"Auto-created enrollment {enrollment_id} for lead {lead_id}")
        except Exception as e:
            logger.error(f"Failed to auto-create enrollment for lead {lead_id}: {str(e)}")

    # Create audit log if there are changes
    if changes:
        audit = AuditLog(
            lead_id=lead_id,
            user_id=current_user["user_id"],
            user_email=current_user["email"],
            user_name=current_user["full_name"],
            action=AuditAction.UPDATED,
            changes=changes
        )
        await audit.insert()

    logger.info(f"Lead updated: {lead_id} by {current_user['email']}")

    return lead_to_response(lead)


@router.get("/{lead_id}/comments")
async def get_comments(
    lead_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Get all comments for a lead
    """
    lead = await Lead.find_one(Lead.lead_id == lead_id, Lead.is_deleted == False)

    if not lead:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lead not found"
        )

    # Agent restriction: only view comments for leads assigned or reassigned to them
    if current_user["role"] == "agent":
        user_id = current_user["user_id"]
        if lead.assigned_to != user_id and lead.reassign_to != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have permission to view comments for this lead"
            )

    return {
        "lead_id": lead_id,
        "comments": lead.comments or []
    }


@router.post("/{lead_id}/comments")
async def add_comment(
    lead_id: str,
    request: CommentCreateRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Add a comment to a lead
    Agents can only add comments to leads assigned/reassigned to them
    Comments are permanent (no edit/delete)
    """
    lead = await Lead.find_one(Lead.lead_id == lead_id, Lead.is_deleted == False)

    if not lead:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lead not found"
        )

    # Agent restriction: only add comments to leads assigned or reassigned to them
    if current_user["role"] == "agent":
        user_id = current_user["user_id"]
        if lead.assigned_to != user_id and lead.reassign_to != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have permission to add comments to this lead"
            )

    # Create new comment
    new_comment = {
        "text": request.text,
        "created_at": datetime.utcnow(),
        "created_by": current_user["user_id"],
        "created_by_name": current_user["full_name"]
    }

    # Add to comments list
    lead.comments.append(new_comment)
    lead.updated_at = datetime.utcnow()
    lead.last_modified_by = current_user["user_id"]

    await lead.save()

    # Create audit log
    audit = AuditLog(
        lead_id=lead_id,
        user_id=current_user["user_id"],
        user_email=current_user["email"],
        user_name=current_user["full_name"],
        action=AuditAction.UPDATED,
        changes=[{"field": "comments", "old_value": None, "new_value": f"Added comment: {request.text[:50]}..."}]
    )
    await audit.insert()

    logger.info(f"Comment added to lead {lead_id} by {current_user['email']}")

    return {
        "message": "Comment added successfully",
        "comment": new_comment
    }


@router.delete("/{lead_id}")
async def delete_lead(
    lead_id: str,
    current_user: dict = Depends(get_current_admin)
):
    """
    Soft delete a lead (Admin only)
    """
    logger.debug(f"Deleting lead {lead_id} by user={current_user['email']}")
    lead = await Lead.find_one(Lead.lead_id == lead_id, Lead.is_deleted == False)

    if not lead:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lead not found"
        )

    # Soft delete
    lead.is_deleted = True
    lead.updated_at = datetime.utcnow()
    lead.last_modified_by = current_user["user_id"]
    await lead.save()

    # Create audit log
    audit = AuditLog(
        lead_id=lead_id,
        user_id=current_user["user_id"],
        user_email=current_user["email"],
        user_name=current_user["full_name"],
        action=AuditAction.DELETED,
        changes=[{"field": "is_deleted", "old_value": False, "new_value": True}]
    )
    await audit.insert()

    logger.info(f"Lead deleted: {lead_id} by {current_user['email']}")

    return {"message": "Lead deleted successfully"}


@router.get("/{lead_id}/audit")
async def get_lead_audit_trail(
    lead_id: str,
    current_user: dict = Depends(get_current_admin)
):
    """
    Get audit trail for a lead (Admin only)
    """
    lead = await Lead.find_one(Lead.lead_id == lead_id)

    if not lead:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lead not found"
        )

    audit_logs = await AuditLog.find(
        AuditLog.lead_id == lead_id
    ).sort("-timestamp").to_list()

    return {
        "lead_id": lead_id,
        "audit_trail": [
            {
                "id": str(log.id),
                "user_email": log.user_email,
                "user_name": log.user_name,
                "action": log.action,
                "changes": log.changes,
                "timestamp": log.timestamp
            }
            for log in audit_logs
        ]
    }
