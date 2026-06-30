"""
Enrollment Management Routes
"""
# Updated: Bug fixes for bulk upload
from fastapi import APIRouter, HTTPException, status, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, date, timedelta, timezone
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.schemas.enrollment import (
    EnrollmentCreateRequest,
    EnrollmentUpdateRequest,
    EnrollmentResponse,
    EnrollmentListResponse,
    EnrollmentStatsResponse,
    FollowUpCreateRequest,
    BulkUploadResponse
)
from app.models.enrollment import Enrollment, ConnectStatus, ActionTaken
from app.models.lead import Trimester, ServicePartner, ServiceEnrolled, Lead, LeadStatus
from app.utils.lead_id import generate_lead_id
from app.models.audit_log import AuditLog, AuditAction
from app.models.enrollment_audit_log import EnrollmentAuditLog, EnrollmentAuditAction
from app.models.user import User
from app.middleware.auth_middleware import get_current_user, get_current_admin, get_current_super_admin
from app.database import get_database
from app.services.journey_service import build_journey_for_service
from app.services.enrollment_helpers import reinstantiate_care_journey
from app.services.journey_ops import compute_care_triggers, stop_journey_steps
from pydantic import BaseModel
import logging
import math
import re
import uuid

logger = logging.getLogger(__name__)

router = APIRouter()


async def generate_enrollment_id() -> str:
    """Generate unique enrollment ID: ENR_DDMMYYYY_XXX"""
    today = datetime.now()
    date_str = today.strftime("%d%m%Y")
    prefix = f"ENR_{date_str}_"

    # Find the highest number for today using regex
    existing = await Enrollment.find(
        {"enrollment_id": {"$regex": f"^{prefix}"}}
    ).sort("-enrollment_id").first_or_none()

    if existing:
        # Extract the number part and increment
        try:
            last_num = int(existing.enrollment_id.split("_")[-1])
            new_num = last_num + 1
        except (ValueError, IndexError):
            new_num = 1
    else:
        new_num = 1

    return f"{prefix}{new_num:03d}"


def enrollment_to_response(enrollment: Enrollment) -> dict:
    """Convert Enrollment document to response dict"""
    return {
        "id": str(enrollment.id),
        "enrollment_id": enrollment.enrollment_id,
        "linked_lead_id": enrollment.linked_lead_id,

        # Timestamps
        "created_at": enrollment.created_at,
        "updated_at": enrollment.updated_at,

        # Billing Info
        "billed_date": enrollment.billed_date,
        "package_billed": enrollment.package_billed,

        # HCLH Details
        "hclhc_spoc": enrollment.hclhc_spoc,
        "hcl_facility": enrollment.hcl_facility,

        # User Details
        "uhid": enrollment.uhid,
        "subscriber_name": enrollment.subscriber_name,
        "dob": enrollment.dob,
        "employee_id": enrollment.employee_id,
        "name": enrollment.name,
        "phone_number": enrollment.phone_number,
        "email": enrollment.email,
        "address": enrollment.address,

        # Service Details
        "trimester": enrollment.trimester.value if enrollment.trimester else None,
        "service_enrolled": enrollment.service_enrolled if enrollment.service_enrolled else None,
        "package_name_enrolled": enrollment.package_name_enrolled,
        "doctor_name": enrollment.doctor_name,
        "service_partner": enrollment.service_partner if enrollment.service_partner else None,
        "partner_centre_selected": enrollment.partner_centre_selected,
        "partner_gynaecologist": enrollment.partner_gynaecologist,

        # Status
        "connect_status": enrollment.connect_status.value if enrollment.connect_status else None,
        "action_taken": enrollment.action_taken.value if enrollment.action_taken else None,

        # Follow-up Tracking
        "follow_up_date": enrollment.follow_up_date,
        "next_follow_up_date": enrollment.next_follow_up_date,
        "customer_feedback": enrollment.customer_feedback,
        "remarks": enrollment.remarks,

        # Follow-ups History
        "follow_ups": enrollment.follow_ups,

        # Care Journey (snapshot of the service's journey template)
        "journey": getattr(enrollment, "journey", []) or [],
        "journey_status": getattr(enrollment, "journey_status", "active"),
        "journey_stopped_reason": getattr(enrollment, "journey_stopped_reason", None),
        "journey_stopped_by_name": getattr(enrollment, "journey_stopped_by_name", None),
        "journey_stopped_at": getattr(enrollment, "journey_stopped_at", None),
        "journey_flag": getattr(enrollment, "journey_flag", None),
        "journey_flag_note": getattr(enrollment, "journey_flag_note", None),
        "journey_classification": getattr(enrollment, "journey_classification", None),
        "converted_to_lead_id": getattr(enrollment, "converted_to_lead_id", None),
        "do_not_contact": bool(getattr(enrollment, "do_not_contact", False)),
        "dnc_reason": getattr(enrollment, "dnc_reason", None),
        # Agent-facing trigger hints (spec 4e).
        "journey_triggers": compute_care_triggers(
            enrollment.service_enrolled, enrollment.trimester
        ),

        # Assignment
        "assigned_to": enrollment.assigned_to,
        "assigned_to_name": enrollment.assigned_to_name,
        "assigned_date": enrollment.assigned_date,
        "reassigned_to": enrollment.reassigned_to,
        "reassigned_to_name": enrollment.reassigned_to_name,
        "reassigned_date": enrollment.reassigned_date,

        # System
        "created_by": enrollment.created_by,
        "created_by_name": enrollment.created_by_name,
    }


@router.get("/stats", response_model=EnrollmentStatsResponse)
async def get_enrollment_stats(
    current_user: dict = Depends(get_current_user)
):
    """Get enrollment statistics by partner and status"""
    # MongoDB stores all dates in UTC
    # Server runs in IST (UTC+5:30), so we need to convert IST date boundaries to UTC
    # IST midnight = UTC previous day 18:30
    IST_OFFSET = timedelta(hours=5, minutes=30)

    today = date.today()  # Local date (IST)
    # IST midnight today (00:00:00 IST) = UTC yesterday 18:30:00
    today_start_ist = datetime.combine(today, datetime.min.time())
    today_start_utc = today_start_ist - IST_OFFSET

    # IST end of today (23:59:59 IST) = UTC today 18:29:59
    today_end_ist = datetime.combine(today, datetime.max.time())
    today_end_utc = today_end_ist - IST_OFFSET

    is_agent = current_user.get("role") == "agent"
    agent_name = current_user.get("full_name", "")
    db = get_database()

    logger.info(f"Fetching stats for user: {agent_name}, role: {current_user.get('role')}")
    logger.info(f"Today (IST): {today}, UTC range: {today_start_utc} to {today_end_utc}")

    # Base query - for agents, filter by hclhc_spoc (case-insensitive)
    if is_agent and agent_name:
        # Use case-insensitive regex match for hclhc_spoc
        base_query = {
            "is_deleted": False,
            "hclhc_spoc": {"$regex": f"^{re.escape(agent_name)}$", "$options": "i"}
        }
    else:
        base_query = {"is_deleted": False}

    # 1. Total count (for agents: total where they are HCLHC SPOC)
    total = await db.enrollments.count_documents(base_query)
    logger.info(f"Total enrollments for {agent_name}: {total}")

    # Agent-specific stats
    new_today = 0
    assigned_today = 0
    follow_up_today = 0

    if is_agent and agent_name:
        hclhc_filter = {"$regex": f"^{re.escape(agent_name)}$", "$options": "i"}

        # 2. New Enrollments Today - created today where she is HCLHC SPOC
        new_today = await db.enrollments.count_documents({
            "is_deleted": False,
            "hclhc_spoc": hclhc_filter,
            "created_at": {"$gte": today_start_utc, "$lte": today_end_utc}
        })

        # 3. Enrollments Assigned Today - where hclhc_spoc is this agent AND
        # (assigned_date OR reassigned_date is today) - irrespective of created date
        assigned_today = await db.enrollments.count_documents({
            "is_deleted": False,
            "hclhc_spoc": hclhc_filter,
            "$or": [
                {"assigned_date": {"$gte": today_start_utc, "$lte": today_end_utc}},
                {"reassigned_date": {"$gte": today_start_utc, "$lte": today_end_utc}}
            ]
        })

        # 4. Follow-ups Today - where she is HCLHC SPOC AND next_follow_up_date is today
        follow_up_today = await db.enrollments.count_documents({
            "is_deleted": False,
            "hclhc_spoc": hclhc_filter,
            "next_follow_up_date": {"$gte": today_start_utc, "$lte": today_end_utc}
        })

        logger.info(f"Agent stats - new_today: {new_today}, assigned_today: {assigned_today}, follow_up_today: {follow_up_today}")
    else:
        # Admin stats - new today is just created today
        new_today = await db.enrollments.count_documents({
            **base_query,
            "created_at": {"$gte": today_start_utc, "$lte": today_end_utc}
        })

    # By partner
    by_partner = {}
    for partner in ServicePartner:
        partner_query = {**base_query, "service_partner": partner.value}
        count = await db.enrollments.count_documents(partner_query)
        if count > 0:
            by_partner[partner.value] = count

    # By connect status
    by_status = {}
    for status_val in ConnectStatus:
        status_query = {**base_query, "connect_status": status_val.value}
        count = await db.enrollments.count_documents(status_query)
        if count > 0:
            by_status[status_val.value] = count

    return {
        "total": total,
        "new_today": new_today,
        "assigned_today": assigned_today,
        "follow_up_today": follow_up_today,
        "by_partner": by_partner,
        "by_status": by_status
    }


@router.get("/bulk-upload/template")
async def get_bulk_upload_template(
    current_user: dict = Depends(get_current_user)
):
    """
    Download a CSV template for bulk enrollment upload
    """
    # Define template columns
    columns = [
        "name", "phone_number", "email", "uhid", "employee_id", "subscriber_name", "dob",
        "billed_date", "package_billed", "hclhc_spoc", "hcl_facility",
        "trimester", "service_enrolled", "package_name_enrolled",
        "service_partner", "partner_centre_selected", "partner_gynaecologist",
        "doctor_name", "address", "connect_status", "action_taken",
        "customer_feedback", "remarks", "next_follow_up_date"
    ]

    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    # Add a sample row
    writer.writerow([
        "Jane Doe", "9876543210", "jane@example.com", "UHID002", "EMP002", "John Doe", "1990-01-15",
        "2026-01-15", "Premium Package", "Agent Name", "Delhi Facility",
        "Trimester 2", "Tulip Antenatal", "Premium",
        "Apollo Cradle", "Kondapur Center", "Dr. Gynae",
        "Dr. Smith", "123 Main St, Delhi", "Connected", "Appointment Booked",
        "Good feedback", "First call completed", "2026-02-15"
    ])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=enrollments_bulk_upload_template.csv"}
    )


@router.post("/bulk-upload", response_model=BulkUploadResponse)
async def bulk_upload_enrollments(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_admin)
):
    """Bulk upload enrollments from CSV file (Admin only)"""
    if not file.filename.endswith('.csv'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV files are supported"
        )

    errors = []
    created_count = 0
    total_rows = 0

    try:
        # Excel-exported CSVs are commonly cp1252/latin-1, not utf-8 —
        # try a few encodings before failing.
        contents = await file.read()
        decoded = None
        for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
            try:
                decoded = contents.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Could not decode file. Please save it as CSV UTF-8 from Excel."
            )
        reader = csv.DictReader(io.StringIO(decoded))

        for row_num, row in enumerate(reader, start=2):
            total_rows += 1
            try:
                # Optional field - Subscriber Name
                subscriber_name = row.get('Subscriber Name', row.get('subscriber_name', '')).strip() or None

                # At least one identifier required: Email, UHID, or Contact No.
                email = row.get('Email', row.get('email', '')).strip() or None
                uhid = row.get('UHID', row.get('uhid', '')).strip() or None
                phone = row.get('Contact No.', row.get('Phone Number', row.get('phone_number', ''))).strip() or None
                employee_id_val = row.get('EmployeeID', row.get('employee_id', '')).strip() or None

                if not email and not uhid and not phone:
                    errors.append({"row": row_num, "error": "At least one of Email, UHID, or Contact No. is required"})
                    continue

                # Validate phone number format if provided
                if phone:
                    if len(phone) != 10 or not phone.isdigit():
                        errors.append({"row": row_num, "error": f"Invalid Contact No.: {phone} (must be 10 digits)"})
                        continue
                    if phone[0] not in "6789":
                        errors.append({"row": row_num, "error": f"Invalid Contact No.: {phone} (must start with 6, 7, 8, or 9)"})
                        continue

                # Parse service partner
                partner_str = row.get('Service (Partner)', row.get('service_partner', '')).strip()
                service_partner = None
                if partner_str:
                    partner_map = {
                        'motherhood': ServicePartner.MOTHERHOOD,
                        'rainbow': ServicePartner.RAINBOW,
                        'fortis': ServicePartner.FORTIS,
                        'apollo cradle': ServicePartner.APOLLO_CRADLE,
                        'cloud 9': ServicePartner.CLOUD_9,
                        'hcl healthcare': ServicePartner.HCL_HEALTHCARE,
                        'mamily': ServicePartner.MAMILY,
                        'others': ServicePartner.OTHERS,
                    }
                    service_partner = partner_map.get(partner_str.lower(), ServicePartner.OTHERS)

                # Parse connect status
                status_str = row.get('Connect Status', row.get('connect_status', '')).strip()
                connect_status = None
                if status_str:
                    status_map = {
                        'connected': ConnectStatus.CONNECTED,
                        'no response': ConnectStatus.NO_RESPONSE,
                        'follow up required': ConnectStatus.FOLLOW_UP_REQUIRED,
                        'others': ConnectStatus.OTHERS,
                    }
                    connect_status = status_map.get(status_str.lower())

                # Parse action taken
                action_str = row.get('Action Taken', row.get('action_taken', '')).strip()
                action_taken = None
                if action_str:
                    action_map = {
                        'appointment booked': ActionTaken.APPOINTMENT_BOOKED,
                        'feedback taken': ActionTaken.FEEDBACK_TAKEN,
                        'no action required': ActionTaken.NO_ACTION_REQUIRED,
                        'liasoned with partner team': ActionTaken.LIASONED_WITH_PARTNER,
                    }
                    action_taken = action_map.get(action_str.lower())

                # Parse trimester
                trimester_str = row.get('Current Trimester', row.get('trimester', '')).strip()
                trimester = None
                if trimester_str:
                    trimester_map = {
                        'trimester 1': Trimester.TRIMESTER_1,
                        'trimester1': Trimester.TRIMESTER_1,
                        '1': Trimester.TRIMESTER_1,
                        'trimester 2': Trimester.TRIMESTER_2,
                        'trimester2': Trimester.TRIMESTER_2,
                        '2': Trimester.TRIMESTER_2,
                        'trimester 3': Trimester.TRIMESTER_3,
                        'trimester3': Trimester.TRIMESTER_3,
                        '3': Trimester.TRIMESTER_3,
                        'not conceived': Trimester.NOT_CONCEIVED,
                    }
                    trimester = trimester_map.get(trimester_str.lower())

                # Parse service enrolled
                service_enrolled_str = row.get('Service Enrolled', row.get('service_enrolled', '')).strip()
                service_enrolled = None
                if service_enrolled_str:
                    service_enrolled_map = {
                        'preconception': ServiceEnrolled.PRE_CONCEPTION,
                        'pre conception': ServiceEnrolled.PRE_CONCEPTION,
                        'pre-conception': ServiceEnrolled.PRE_CONCEPTION,
                        'antenatal': ServiceEnrolled.ANTENATAL,
                        'maternitywellness': ServiceEnrolled.MATERNITY_WELLNESS,
                        'maternity wellness': ServiceEnrolled.MATERNITY_WELLNESS,
                        'maternity-wellness': ServiceEnrolled.MATERNITY_WELLNESS,
                    }
                    service_enrolled = service_enrolled_map.get(service_enrolled_str.lower())

                # Get package name enrolled
                package_name_enrolled = row.get('Package Name Enrolled', row.get('package_name_enrolled', '')).strip() or None

                # Parse dates
                billed_date = None
                billed_str = row.get('Billed Date', row.get('billed_date', '')).strip()
                if billed_str:
                    try:
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                            try:
                                billed_date = datetime.strptime(billed_str, fmt).date()
                                break
                            except ValueError:
                                continue
                    except:
                        pass

                dob = None
                dob_str = row.get('DOB', row.get('dob', '')).strip()
                if dob_str:
                    try:
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                            try:
                                dob = datetime.strptime(dob_str, fmt).date()
                                break
                            except ValueError:
                                continue
                    except:
                        pass

                follow_up_date = None
                follow_str = row.get('Follow Up Date', row.get('follow_up_date', '')).strip()
                if follow_str:
                    try:
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                            try:
                                follow_up_date = datetime.strptime(follow_str, fmt)
                                break
                            except ValueError:
                                continue
                    except:
                        pass

                next_follow_up_date = None
                next_follow_str = row.get('Next Follow Up Date', row.get('next_follow_up_date', '')).strip()
                if next_follow_str:
                    try:
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                            try:
                                next_follow_up_date = datetime.strptime(next_follow_str, fmt)
                                break
                            except ValueError:
                                continue
                    except:
                        pass

                # Create enrollment
                enrollment = Enrollment(
                    enrollment_id=await generate_enrollment_id(),
                    subscriber_name=subscriber_name,
                    employee_id=employee_id_val,
                    phone_number=phone,
                    email=email,
                    billed_date=billed_date,
                    package_billed=row.get('Package Billed', row.get('package_billed', '')).strip() or None,
                    hclhc_spoc=row.get('HCLH SPOC', row.get('hclhc_spoc', '')).strip() or None,
                    hcl_facility=row.get('HCL Facility', row.get('hcl_facility', '')).strip() or None,
                    uhid=uhid,
                    dob=dob,
                    name=row.get('Name', row.get('name', '')).strip() or None,
                    address=row.get('Address', row.get('address', '')).strip() or None,
                    trimester=trimester,
                    service_enrolled=service_enrolled,
                    package_name_enrolled=package_name_enrolled,
                    doctor_name=row.get('Doctor Name', row.get('doctor_name', '')).strip() or None,
                    service_partner=service_partner,
                    partner_centre_selected=row.get('Partner Centre Selected', row.get('partner_centre_selected', '')).strip() or None,
                    partner_gynaecologist=row.get('Partner Gynaecologist', row.get('partner_gynaecologist', '')).strip() or None,
                    connect_status=connect_status,
                    action_taken=action_taken,
                    follow_up_date=follow_up_date,
                    next_follow_up_date=next_follow_up_date,
                    customer_feedback=row.get('Customer Feedback', row.get('customer_feedback', '')).strip() or None,
                    remarks=row.get('Remarks', row.get('remarks', '')).strip() or None,
                    created_by=current_user["user_id"],
                    created_by_name=current_user.get("full_name", current_user["email"]),
                    assigned_to=current_user["user_id"],
                    assigned_to_name=current_user.get("full_name", current_user["email"]),
                )
                await enrollment.insert()
                created_count += 1

            except Exception as e:
                logger.error(f"Row {row_num} bulk upload error: {type(e).__name__}: {str(e)}")
                errors.append({"row": row_num, "error": str(e)})

        logger.info(f"Bulk upload by {current_user['email']}: {created_count}/{total_rows} created")

        return {
            "success": True,
            "message": f"Processed {total_rows} rows. Created {created_count} enrollments. {len(errors)} errors.",
            "total_rows": total_rows,
            "created": created_count,
            "errors": errors
        }

    except Exception as e:
        logger.error(f"Bulk upload error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to process file: {str(e)}"
        )


@router.get("/export/excel")
async def export_enrollments_excel(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD), inclusive, IST"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD), inclusive, IST"),
    current_user: dict = Depends(get_current_user)
):
    """
    Export enrollments to Excel.
    Admins/super-admins export all enrollments; agents export only enrollments
    where they are the HCLHC SPOC. Optional created_at date range filter (IST).
    """
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Enrollments
    ws = wb.active
    ws.title = "Enrollments"

    headers = [
        "Enrollment ID", "Linked Lead", "Billed Date", "Package Billed",
        "HCLH SPOC", "HCL Facility", "UHID", "Subscriber Name", "DOB",
        "EmployeeID", "Name", "Contact No.", "Email", "Address",
        "Current Trimester", "Service Enrolled", "Package Name Enrolled", "Doctor Name",
        "Service (Partner)", "Partner Centre Selected", "Partner Gynaecologist",
        "Connect Status", "Action Taken", "Follow Up Date", "Next Follow Up Date",
        "Customer Feedback", "Remarks", "Assigned To", "Created At"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Build query with optional IST-aware created_at date range.
    # created_at is stored in UTC; the picker sends IST calendar dates, so we
    # offset by IST (+5:30) to get the correct UTC boundaries.
    IST_OFFSET = timedelta(hours=5, minutes=30)
    query: dict = {"is_deleted": False}
    # Agents can only export enrollments where they are the HCLHC SPOC
    if current_user.get("role") == "agent":
        user_name = current_user.get("full_name", "")
        query["hclhc_spoc"] = {"$regex": f"^{re.escape(user_name)}$", "$options": "i"}
    created_range: dict = {}
    if start_date:
        try:
            created_range["$gte"] = datetime.strptime(start_date, "%Y-%m-%d") - IST_OFFSET
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid start_date (expected YYYY-MM-DD)")
    if end_date:
        try:
            created_range["$lt"] = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - IST_OFFSET
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid end_date (expected YYYY-MM-DD)")
    if created_range:
        query["created_at"] = created_range

    enrollments = await Enrollment.find(query).sort("-created_at").to_list()

    for row_num, enrollment in enumerate(enrollments, 2):
        row_data = [
            enrollment.enrollment_id,
            enrollment.linked_lead_id,
            str(enrollment.billed_date) if enrollment.billed_date else None,
            enrollment.package_billed,
            enrollment.hclhc_spoc,
            enrollment.hcl_facility,
            enrollment.uhid,
            enrollment.subscriber_name,
            str(enrollment.dob) if enrollment.dob else None,
            enrollment.employee_id,
            enrollment.name,
            enrollment.phone_number,
            enrollment.email,
            enrollment.address,
            enrollment.trimester.value if enrollment.trimester else None,
            enrollment.service_enrolled if enrollment.service_enrolled else None,
            enrollment.package_name_enrolled,
            enrollment.doctor_name,
            enrollment.service_partner if enrollment.service_partner else None,
            enrollment.partner_centre_selected,
            enrollment.partner_gynaecologist,
            enrollment.connect_status.value if enrollment.connect_status else None,
            enrollment.action_taken.value if enrollment.action_taken else None,
            enrollment.follow_up_date.strftime("%Y-%m-%d %H:%M") if enrollment.follow_up_date else None,
            enrollment.next_follow_up_date.strftime("%Y-%m-%d %H:%M") if enrollment.next_follow_up_date else None,
            enrollment.customer_feedback,
            enrollment.remarks,
            enrollment.assigned_to_name,
            enrollment.created_at.strftime("%Y-%m-%d %H:%M") if enrollment.created_at else None,
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Sheet 2: Follow-ups
    ws_followups = wb.create_sheet("Follow-ups")
    followup_headers = ["Enrollment ID", "Subscriber Name", "Follow-up #", "Date", "Connect Status", "Action Taken", "Feedback", "Remarks", "Created By"]

    for col, header in enumerate(followup_headers, 1):
        cell = ws_followups.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    followup_row = 2
    for enrollment in enrollments:
        for followup in enrollment.follow_ups:
            ws_followups.cell(row=followup_row, column=1, value=enrollment.enrollment_id).border = thin_border
            ws_followups.cell(row=followup_row, column=2, value=enrollment.subscriber_name).border = thin_border
            ws_followups.cell(row=followup_row, column=3, value=followup.get('follow_up_number', '')).border = thin_border
            ws_followups.cell(row=followup_row, column=4, value=followup.get('date', '')).border = thin_border
            ws_followups.cell(row=followup_row, column=5, value=followup.get('connect_status', '')).border = thin_border
            ws_followups.cell(row=followup_row, column=6, value=followup.get('action_taken', '')).border = thin_border
            ws_followups.cell(row=followup_row, column=7, value=followup.get('feedback', '')).border = thin_border
            ws_followups.cell(row=followup_row, column=8, value=followup.get('remarks', '')).border = thin_border
            ws_followups.cell(row=followup_row, column=9, value=followup.get('created_by_name', '')).border = thin_border
            followup_row += 1

    # Auto-adjust follow-ups sheet
    for col in ws_followups.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_followups.column_dimensions[column].width = min(max_length + 2, 50)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"enrollments_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("", response_model=EnrollmentListResponse)
async def get_enrollments(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=100),
    search: Optional[str] = None,
    connect_status: Optional[List[str]] = Query(None),
    action_taken: Optional[List[str]] = Query(None),
    service_partner: Optional[List[str]] = Query(None),
    service_enrolled: Optional[List[str]] = Query(None),
    package: Optional[str] = None,
    uhid: Optional[List[str]] = Query(None),
    hclhc_spoc: Optional[str] = None,
    created_date_from: Optional[str] = None,
    created_date_to: Optional[str] = None,
    next_follow_up_date: Optional[str] = None,
    assigned_today: Optional[bool] = None,
    my_role: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get enrollments with pagination and filters"""
    try:
        query = {"is_deleted": False}

        # Visibility for agents: an enrollment shows on their page if they are the
        # current follow-up SPOC OR they are the agent who enrolled it (monitor view).
        # `my_role` lets the agent split the two:
        #   'following_up' = mine to act on (I'm the SPOC)
        #   'enrolled'     = I enrolled it but someone else now follows up (read-only)
        agent_or = None
        if current_user.get("role") == "agent":
            user_name = current_user.get("full_name", "")
            user_id = current_user["user_id"]
            spoc_regex = {"$regex": f"^{re.escape(user_name)}$", "$options": "i"}
            spoc_match = {"hclhc_spoc": spoc_regex}
            enrolled_monitor = {"$and": [
                {"created_by": user_id},
                {"hclhc_spoc": {"$not": spoc_regex}},
            ]}
            if my_role == "following_up":
                agent_or = [spoc_match]
            elif my_role == "enrolled":
                agent_or = [enrolled_monitor]
            else:
                agent_or = [spoc_match, {"created_by": user_id}]

        # Apply filters (support multiple values with $in)
        if connect_status and len(connect_status) > 0:
            query["connect_status"] = {"$in": connect_status}
        if action_taken and len(action_taken) > 0:
            query["action_taken"] = {"$in": action_taken}
        if service_partner and len(service_partner) > 0:
            query["service_partner"] = {"$in": service_partner}
        # Service enrolled (standardized field) - multi-select
        if service_enrolled and len(service_enrolled) > 0:
            svc_alt = "|".join(re.escape(s.strip()) for s in service_enrolled if s and s.strip())
            if svc_alt:
                query["service_enrolled"] = {"$regex": f"^\\s*({svc_alt})\\s*$", "$options": "i"}
        # Package (free text) - partial, case-insensitive match on package_name_enrolled
        if package and package.strip():
            query["package_name_enrolled"] = {"$regex": re.escape(package.strip()), "$options": "i"}
        # UHID: whitespace-tolerant case-insensitive match (some rows have padding)
        if uhid and len(uhid) > 0:
            uhid_alternation = "|".join(re.escape(u.strip()) for u in uhid if u and u.strip())
            if uhid_alternation:
                query["uhid"] = {"$regex": f"^\\s*({uhid_alternation})\\s*$", "$options": "i"}
        if hclhc_spoc:
            query["hclhc_spoc"] = {"$regex": re.escape(hclhc_spoc), "$options": "i"}

        # Date range filter for created_at
        if created_date_from or created_date_to:
            created_at_filter = {}
            if created_date_from:
                try:
                    from_date = datetime.strptime(created_date_from, "%Y-%m-%d").date()
                    created_at_filter["$gte"] = datetime.combine(from_date, datetime.min.time())
                except ValueError:
                    pass
            if created_date_to:
                try:
                    to_date = datetime.strptime(created_date_to, "%Y-%m-%d").date()
                    created_at_filter["$lte"] = datetime.combine(to_date, datetime.max.time())
                except ValueError:
                    pass
            if created_at_filter:
                query["created_at"] = created_at_filter

        if next_follow_up_date:
            try:
                filter_date = datetime.strptime(next_follow_up_date, "%Y-%m-%d").date()
                day_start = datetime.combine(filter_date, datetime.min.time())
                day_end = datetime.combine(filter_date, datetime.max.time())
                query["next_follow_up_date"] = {"$gte": day_start, "$lte": day_end}
            except ValueError:
                pass

        # Search across all common identifying fields (escape regex special chars for security)
        search_conditions = None
        if search and search.strip():
            escaped_search = re.escape(search.strip())
            search_conditions = [
                {"subscriber_name": {"$regex": escaped_search, "$options": "i"}},
                {"employee_id": {"$regex": escaped_search, "$options": "i"}},
                {"name": {"$regex": escaped_search, "$options": "i"}},
                {"phone_number": {"$regex": escaped_search}},
                {"enrollment_id": {"$regex": escaped_search, "$options": "i"}},
                {"linked_lead_id": {"$regex": escaped_search, "$options": "i"}},
                {"email": {"$regex": escaped_search, "$options": "i"}},
                {"uhid": {"$regex": escaped_search, "$options": "i"}},
                {"package_name_enrolled": {"$regex": escaped_search, "$options": "i"}},
                {"doctor_name": {"$regex": escaped_search, "$options": "i"}},
                {"hclhc_spoc": {"$regex": escaped_search, "$options": "i"}},
                {"partner_centre_selected": {"$regex": escaped_search, "$options": "i"}},
                {"assigned_to_name": {"$regex": escaped_search, "$options": "i"}},
                {"reassign_to_name": {"$regex": escaped_search, "$options": "i"}},
                {"created_by_name": {"$regex": escaped_search, "$options": "i"}},
            ]

        # Combine agent visibility and search as ANDed $or groups (each must hold)
        and_groups = []
        if agent_or is not None:
            and_groups.append({"$or": agent_or})
        if search_conditions is not None:
            and_groups.append({"$or": search_conditions})
        if and_groups:
            query.setdefault("$and", []).extend(and_groups)

        # "Assigned today" quick filter: assigned_date OR reassigned_date is today (IST).
        # Mirrors the Assigned-Today KPI card. Applied last as an $and wrapper.
        if assigned_today:
            IST_OFFSET = timedelta(hours=5, minutes=30)
            t = date.today()
            a_start = datetime.combine(t, datetime.min.time()) - IST_OFFSET
            a_end = datetime.combine(t, datetime.max.time()) - IST_OFFSET
            query = {"$and": [query, {"$or": [
                {"assigned_date": {"$gte": a_start, "$lte": a_end}},
                {"reassigned_date": {"$gte": a_start, "$lte": a_end}},
            ]}]}

        # Get total count
        total = await Enrollment.find(query).count()
        pages = math.ceil(total / per_page) if total > 0 else 1

        # Pagination
        skip = (page - 1) * per_page
        enrollments = await Enrollment.find(query).sort("-created_at").skip(skip).limit(per_page).to_list()

        return {
            "enrollments": [enrollment_to_response(e) for e in enrollments],
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": pages
        }
    except Exception as e:
        logger.error(f"Error in get_enrollments: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.post("", response_model=EnrollmentResponse)
async def create_enrollment(
    enrollment_data: EnrollmentCreateRequest,
    current_user: dict = Depends(get_current_user)
):
    """Create a new enrollment"""
    logger.info(f"Creating enrollment by user: {current_user.get('full_name')} ({current_user.get('email')})")

    # Exclude assigned_to from model_dump to avoid duplicate kwarg
    data_dict = enrollment_data.model_dump(exclude_unset=True, exclude={'assigned_to'})

    logger.debug(f"Initial hclhc_spoc from request: '{data_dict.get('hclhc_spoc')}'")

    # Handle hclhc_spoc - if empty/None/whitespace, default to current user's name
    hclhc_spoc_value = data_dict.get('hclhc_spoc')
    if not hclhc_spoc_value or (isinstance(hclhc_spoc_value, str) and not hclhc_spoc_value.strip()):
        hclhc_spoc = current_user.get("full_name") or current_user.get("email")
        logger.info(f"hclhc_spoc was empty, defaulting to current user: '{hclhc_spoc}'")
    else:
        hclhc_spoc = hclhc_spoc_value.strip()
        logger.info(f"Using provided hclhc_spoc: '{hclhc_spoc}'")

    # Remove hclhc_spoc from data_dict to avoid conflict, we'll set it explicitly
    data_dict.pop('hclhc_spoc', None)

    # Sync assigned_to and assigned_to_name from hclhc_spoc
    assigned_to = current_user["user_id"]
    assigned_to_name = hclhc_spoc

    # Look up user by full_name matching hclhc_spoc
    spoc_user = await User.find_one({"full_name": {"$regex": f"^{re.escape(hclhc_spoc)}$", "$options": "i"}})
    if spoc_user:
        assigned_to = str(spoc_user.id)
        assigned_to_name = spoc_user.full_name

    enrollment = Enrollment(
        enrollment_id=await generate_enrollment_id(),
        **data_dict,
        hclhc_spoc=hclhc_spoc,  # Set explicitly
        created_by=current_user["user_id"],
        created_by_name=current_user.get("full_name", current_user["email"]),
        assigned_to=assigned_to,
        assigned_to_name=assigned_to_name,
        assigned_date=datetime.utcnow(),  # Set assigned date on creation
    )

    logger.info(f"Enrollment object hclhc_spoc before insert: '{enrollment.hclhc_spoc}'")
    await enrollment.insert()
    logger.info(f"Enrollment {enrollment.enrollment_id} created with hclhc_spoc: '{enrollment.hclhc_spoc}'")

    # Create audit log for enrollment creation
    audit = EnrollmentAuditLog(
        enrollment_id=enrollment.enrollment_id,
        user_id=current_user["user_id"],
        user_email=current_user["email"],
        user_name=current_user.get("full_name", current_user["email"]),
        action=EnrollmentAuditAction.CREATED,
        changes=[{"field": "enrollment", "old_value": None, "new_value": "created"}]
    )
    await audit.insert()

    logger.info(f"Enrollment {enrollment.enrollment_id} created by {current_user['email']}")

    return enrollment_to_response(enrollment)


@router.post("/backfill-spoc")
async def backfill_hclhc_spoc(current_user: dict = Depends(get_current_super_admin)):
    """
    One-time data fix (Super Admin): fill the HCLHC SPOC on old enrollments that
    were created before SPOC was mandatory.

    For every enrollment whose hclhc_spoc is null / empty / missing, look up the
    enrolling agent (created_by) and set hclhc_spoc = that user's full_name. Also
    fill assigned_to / assigned_to_name from the same user IF they are blank.

    Never overwrites an enrollment that already has a non-empty hclhc_spoc.
    Returns counts of what was changed.
    """
    # Defined as a literal path BEFORE the dynamic /{enrollment_id} routes so it
    # is matched as an action, not as an enrollment id.
    blank_spoc_query = {
        "is_deleted": False,
        "$or": [
            {"hclhc_spoc": None},
            {"hclhc_spoc": ""},
            {"hclhc_spoc": {"$exists": False}},
        ],
    }
    candidates = await Enrollment.find(blank_spoc_query).to_list()

    user_cache: dict = {}          # created_by id -> full_name (or None if not found)
    updated = 0
    skipped_no_creator = 0
    assigned_filled = 0

    for enr in candidates:
        # Safety: never touch one that already has a non-empty SPOC.
        if enr.hclhc_spoc and enr.hclhc_spoc.strip():
            continue

        creator_id = enr.created_by
        if not creator_id:
            skipped_no_creator += 1
            continue

        if creator_id in user_cache:
            full_name = user_cache[creator_id]
        else:
            full_name = None
            try:
                u = await User.get(creator_id)
                if u:
                    full_name = u.full_name
            except Exception:
                full_name = None
            user_cache[creator_id] = full_name

        if not full_name:
            skipped_no_creator += 1
            continue

        enr.hclhc_spoc = full_name

        # Fill assignment only when blank — don't disturb an existing owner.
        if not enr.assigned_to:
            enr.assigned_to = creator_id
            assigned_filled += 1
        if not enr.assigned_to_name:
            enr.assigned_to_name = full_name

        enr.updated_at = datetime.utcnow()
        enr.last_modified_by = current_user["user_id"]
        await enr.save()
        updated += 1

    logger.info(
        f"backfill-spoc by {current_user['email']}: "
        f"checked={len(candidates)}, updated={updated}, "
        f"assigned_filled={assigned_filled}, skipped_no_creator={skipped_no_creator}"
    )
    return {
        "message": f"Backfilled HCLHC SPOC on {updated} enrollment(s)",
        "checked": len(candidates),
        "updated": updated,
        "assigned_filled": assigned_filled,
        "skipped_no_creator": skipped_no_creator,
    }


@router.get("/{enrollment_id}", response_model=EnrollmentResponse)
async def get_enrollment(
    enrollment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a single enrollment by ID"""
    enrollment = await Enrollment.find_one(
        Enrollment.enrollment_id == enrollment_id,
        Enrollment.is_deleted == False
    )
    if not enrollment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Enrollment not found"
        )

    # Agents can view enrollments where they are HCLHC SPOC — or, for legacy
    # blank-SPOC enrollments, where they are the agent who enrolled it. Mirrors
    # the edit/follow-up/journey safety net (_agent_can_act).
    if current_user.get("role") == "agent" and not _agent_can_act(enrollment, current_user):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You can only view enrollments where you are the HCLHC SPOC"
        )

    return enrollment_to_response(enrollment)


def _agent_can_act(enrollment: Enrollment, current_user: dict) -> bool:
    """
    Whether a non-admin agent may edit / log follow-ups / work the journey on
    this enrollment.

    True when:
      - the agent is the current HCLHC SPOC, OR
      - (safety net for old enrollments created before SPOC was mandatory) the
        enrollment has no SPOC yet AND the agent is the one who enrolled it
        (created_by == this user).
    """
    user_name = (current_user.get("full_name") or "").strip().lower()
    spoc = (enrollment.hclhc_spoc or "").strip()
    if spoc and spoc.lower() == user_name:
        return True
    if not spoc and enrollment.created_by and enrollment.created_by == current_user.get("user_id"):
        return True
    return False


@router.put("/{enrollment_id}", response_model=EnrollmentResponse)
async def update_enrollment(
    enrollment_id: str,
    update_data: EnrollmentUpdateRequest,
    current_user: dict = Depends(get_current_user)
):
    """Update an enrollment"""
    logger.debug(f"Updating enrollment {enrollment_id} by user={current_user['email']}")
    enrollment = await Enrollment.find_one(
        Enrollment.enrollment_id == enrollment_id,
        Enrollment.is_deleted == False
    )
    if not enrollment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Enrollment not found"
        )

    # Agents can update enrollments where they are HCLHC SPOC — or, for legacy
    # blank-SPOC enrollments, where they are the agent who enrolled it.
    if current_user.get("role") == "agent" and not _agent_can_act(enrollment, current_user):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You can only edit enrollments where you are the HCLHC SPOC"
        )

    # Track changes for audit log
    changes = []
    update_dict = update_data.model_dump(exclude_unset=True)

    for field, new_value in update_dict.items():
        old_value = getattr(enrollment, field, None)

        # Handle enum values for comparison
        if hasattr(old_value, 'value'):
            old_value_str = old_value.value
        else:
            old_value_str = old_value

        if hasattr(new_value, 'value'):
            new_value_str = new_value.value
        else:
            new_value_str = new_value

        # Only record if value actually changed
        if old_value_str != new_value_str:
            changes.append({
                "field": field,
                "old_value": old_value_str,
                "new_value": new_value_str
            })

        setattr(enrollment, field, new_value)

    # Sync assigned_to when hclhc_spoc is updated (this is a reassignment)
    if 'hclhc_spoc' in update_dict and update_dict['hclhc_spoc']:
        hclhc_spoc = update_dict['hclhc_spoc']
        old_assigned_to = enrollment.assigned_to

        spoc_user = await User.find_one({"full_name": {"$regex": f"^{re.escape(hclhc_spoc)}$", "$options": "i"}})
        new_assigned_to = str(spoc_user.id) if spoc_user else None
        new_assigned_to_name = spoc_user.full_name if spoc_user else hclhc_spoc

        # Check if this is a reassignment (assigned_to changed)
        if old_assigned_to and new_assigned_to and old_assigned_to != new_assigned_to:
            # This is a reassignment - set reassigned fields
            enrollment.reassigned_to = new_assigned_to
            enrollment.reassigned_to_name = new_assigned_to_name
            enrollment.reassigned_date = datetime.utcnow()

        # Always update assigned_to to the new user
        if spoc_user:
            enrollment.assigned_to = str(spoc_user.id)
            enrollment.assigned_to_name = spoc_user.full_name
        else:
            enrollment.assigned_to_name = hclhc_spoc

    enrollment.updated_at = datetime.utcnow()
    enrollment.last_modified_by = current_user["user_id"]
    await enrollment.save()

    # Trimester change on a care journey (esp. Antenatal blank/"Not Conceived" ->
    # a real trimester) re-instantiates the journey so the loop materializes.
    # Preserves completed-step attribution + ad-hoc steps. Skips stopped journeys.
    trimester_changed = any(c["field"] == "trimester" for c in changes)
    if trimester_changed and enrollment.journey_status == "active" and not enrollment.do_not_contact:
        try:
            enrollment.journey = await reinstantiate_care_journey(enrollment)
            await enrollment.save()
        except Exception as e:
            logger.error(f"Failed to re-instantiate journey for {enrollment_id}: {e}")

    # Create audit log only if there were changes
    if changes:
        audit = EnrollmentAuditLog(
            enrollment_id=enrollment_id,
            user_id=current_user["user_id"],
            user_email=current_user["email"],
            user_name=current_user.get("full_name", current_user["email"]),
            action=EnrollmentAuditAction.UPDATED,
            changes=changes
        )
        await audit.insert()

    logger.info(f"Enrollment {enrollment_id} updated by {current_user['email']}")

    return enrollment_to_response(enrollment)


@router.delete("/{enrollment_id}")
async def delete_enrollment(
    enrollment_id: str,
    current_user: dict = Depends(get_current_admin)
):
    """Soft delete an enrollment (Admin only)"""
    logger.debug(f"Deleting enrollment {enrollment_id} by user={current_user['email']}")
    enrollment = await Enrollment.find_one(
        Enrollment.enrollment_id == enrollment_id,
        Enrollment.is_deleted == False
    )
    if not enrollment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Enrollment not found"
        )

    enrollment.is_deleted = True
    enrollment.updated_at = datetime.utcnow()
    await enrollment.save()

    # Create audit log for deletion
    audit = EnrollmentAuditLog(
        enrollment_id=enrollment_id,
        user_id=current_user["user_id"],
        user_email=current_user["email"],
        user_name=current_user.get("full_name", current_user["email"]),
        action=EnrollmentAuditAction.DELETED,
        changes=[{"field": "is_deleted", "old_value": False, "new_value": True}]
    )
    await audit.insert()

    logger.info(f"Enrollment {enrollment_id} deleted by {current_user['email']}")

    return {"message": "Enrollment deleted successfully"}


@router.get("/{enrollment_id}/follow-ups")
async def get_follow_ups(
    enrollment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get all follow-ups for an enrollment"""
    enrollment = await Enrollment.find_one(
        Enrollment.enrollment_id == enrollment_id,
        Enrollment.is_deleted == False
    )
    if not enrollment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Enrollment not found"
        )

    return {
        "enrollment_id": enrollment_id,
        "follow_ups": enrollment.follow_ups or []
    }


@router.post("/{enrollment_id}/follow-ups", response_model=EnrollmentResponse)
async def add_follow_up(
    enrollment_id: str,
    follow_up_data: FollowUpCreateRequest,
    current_user: dict = Depends(get_current_user)
):
    """Add a follow-up entry to an enrollment"""
    enrollment = await Enrollment.find_one(
        Enrollment.enrollment_id == enrollment_id,
        Enrollment.is_deleted == False
    )
    if not enrollment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Enrollment not found"
        )

    # Only the current follow-up SPOC (or an admin) may log follow-ups. The agent
    # who enrolled the lead can see it but is read-only once it's handed off — except
    # for legacy blank-SPOC enrollments, where the enroller may still act.
    if current_user.get("role") == "agent" and not _agent_can_act(enrollment, current_user):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only the assigned follow-up SPOC can add follow-ups"
        )

    # Create follow-up entry
    follow_up_number = len(enrollment.follow_ups) + 1
    follow_up_entry = {
        "follow_up_number": follow_up_number,
        "date": datetime.utcnow().isoformat(),
        "connect_status": follow_up_data.connect_status.value if follow_up_data.connect_status else None,
        "action_taken": follow_up_data.action_taken.value if follow_up_data.action_taken else None,
        "feedback": follow_up_data.feedback,
        "remarks": follow_up_data.remarks,
        "next_follow_up_date": follow_up_data.follow_up_date.isoformat() if follow_up_data.follow_up_date else None,
        "created_by": current_user["user_id"],
        "created_by_name": current_user.get("full_name", current_user["email"]),
        "created_at": datetime.utcnow().isoformat()
    }

    enrollment.follow_ups.append(follow_up_entry)

    # Update current status fields
    if follow_up_data.connect_status:
        enrollment.connect_status = follow_up_data.connect_status
    if follow_up_data.action_taken:
        enrollment.action_taken = follow_up_data.action_taken
    if follow_up_data.feedback:
        enrollment.customer_feedback = follow_up_data.feedback
    if follow_up_data.remarks:
        enrollment.remarks = follow_up_data.remarks
    if follow_up_data.follow_up_date:
        enrollment.next_follow_up_date = follow_up_data.follow_up_date

    enrollment.updated_at = datetime.utcnow()
    await enrollment.save()

    # Create audit log for follow-up addition
    follow_up_changes = [
        {"field": f"Follow-up #{follow_up_number}", "old_value": None, "new_value": "Added"}
    ]
    if follow_up_data.connect_status:
        follow_up_changes.append({
            "field": "connect_status",
            "old_value": None,
            "new_value": follow_up_data.connect_status.value
        })
    if follow_up_data.action_taken:
        follow_up_changes.append({
            "field": "action_taken",
            "old_value": None,
            "new_value": follow_up_data.action_taken.value
        })
    if follow_up_data.feedback:
        feedback_summary = follow_up_data.feedback[:50] + "..." if len(follow_up_data.feedback) > 50 else follow_up_data.feedback
        follow_up_changes.append({
            "field": "feedback",
            "old_value": None,
            "new_value": feedback_summary
        })
    if follow_up_data.follow_up_date:
        follow_up_changes.append({
            "field": "next_follow_up_date",
            "old_value": None,
            "new_value": follow_up_data.follow_up_date.isoformat()
        })

    audit = EnrollmentAuditLog(
        enrollment_id=enrollment_id,
        user_id=current_user["user_id"],
        user_email=current_user["email"],
        user_name=current_user.get("full_name", current_user["email"]),
        action=EnrollmentAuditAction.FOLLOW_UP_ADDED,
        changes=follow_up_changes
    )
    await audit.insert()

    logger.info(f"Follow-up #{follow_up_number} added to {enrollment_id} by {current_user['email']}")

    return enrollment_to_response(enrollment)


# ---------------------------------------------------------------------------
# Care Journey execution (worked by the follow-up SPOC / admins)
# ---------------------------------------------------------------------------

class JourneyStepUpdateRequest(BaseModel):
    status: Optional[str] = None            # pending | done | skipped
    planned_date: Optional[datetime] = None
    notes: Optional[str] = None


class JourneyStepCreateRequest(BaseModel):
    name: str
    description: Optional[str] = None
    step_type: Optional[str] = "Other"
    planned_date: Optional[datetime] = None


def _require_spoc_or_admin(enrollment: Enrollment, current_user: dict):
    """
    Only the current follow-up SPOC (or an admin) may work the journey — plus, for
    legacy blank-SPOC enrollments, the agent who enrolled it (see _agent_can_act).
    """
    if current_user.get("role") == "agent" and not _agent_can_act(enrollment, current_user):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only the assigned follow-up SPOC can update the care journey",
        )


async def _get_enrollment_or_404(enrollment_id: str) -> Enrollment:
    enrollment = await Enrollment.find_one(
        Enrollment.enrollment_id == enrollment_id,
        Enrollment.is_deleted == False
    )
    if not enrollment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Enrollment not found"
        )
    return enrollment


@router.post("/{enrollment_id}/journey/instantiate")
async def instantiate_journey(
    enrollment_id: str,
    force: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """
    Build (or rebuild) this enrollment's care journey from the current template
    for its service. Used for enrollments created before journeys existed.
    Skips if a journey already exists unless force=True.
    """
    enrollment = await _get_enrollment_or_404(enrollment_id)
    _require_spoc_or_admin(enrollment, current_user)

    if enrollment.journey and not force:
        return enrollment_to_response(enrollment)

    # Trimester-aware rebuild that preserves done/skipped attribution + ad-hoc steps.
    enrollment.journey = await reinstantiate_care_journey(enrollment)
    enrollment.journey_status = "active"
    enrollment.updated_at = datetime.utcnow()
    await enrollment.save()
    logger.info(
        "instantiate %s: service=%r trimester=%r built=%d",
        enrollment_id,
        enrollment.journey_classification or enrollment.service_enrolled,
        (enrollment.trimester.value if hasattr(enrollment.trimester, "value") else enrollment.trimester),
        len(enrollment.journey or []),
    )
    return enrollment_to_response(enrollment)


@router.put("/{enrollment_id}/journey/{step_id}")
async def update_journey_step(
    enrollment_id: str,
    step_id: str,
    body: JourneyStepUpdateRequest,
    current_user: dict = Depends(get_current_user)
):
    """Update a single journey step: mark done/skipped, reschedule, or add notes."""
    enrollment = await _get_enrollment_or_404(enrollment_id)
    _require_spoc_or_admin(enrollment, current_user)

    journey = enrollment.journey or []
    step = next((s for s in journey if s.get("step_id") == step_id), None)
    if not step:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Journey step not found")

    if body.status is not None:
        if body.status not in ("pending", "done", "skipped"):
            raise HTTPException(status_code=400, detail="status must be pending, done or skipped")
        step["status"] = body.status
        if body.status == "done":
            step["completed_date"] = datetime.utcnow()
            step["completed_by"] = current_user["user_id"]
            step["completed_by_name"] = current_user.get("full_name", current_user["email"])
        else:
            # reverting to pending / skipping clears completion stamp
            step["completed_date"] = None
            step["completed_by"] = None
            step["completed_by_name"] = None
    if body.planned_date is not None:
        step["planned_date"] = body.planned_date
    if body.notes is not None:
        step["notes"] = body.notes

    enrollment.journey = journey
    enrollment.updated_at = datetime.utcnow()
    await enrollment.save()
    return enrollment_to_response(enrollment)


@router.post("/{enrollment_id}/journey")
async def add_journey_step(
    enrollment_id: str,
    body: JourneyStepCreateRequest,
    current_user: dict = Depends(get_current_user)
):
    """Add an ad-hoc step to this one enrollment's journey (does not touch the template)."""
    enrollment = await _get_enrollment_or_404(enrollment_id)
    _require_spoc_or_admin(enrollment, current_user)

    journey = enrollment.journey or []
    max_order = max((s.get("order", 0) for s in journey), default=-1)
    journey.append({
        "step_id": uuid.uuid4().hex[:12],
        "name": body.name.strip(),
        "description": (body.description.strip() if body.description else None),
        "step_type": body.step_type or "Other",
        "planned_date": body.planned_date,
        "status": "pending",
        "completed_date": None,
        "completed_by": None,
        "completed_by_name": None,
        "notes": None,
        "order": max_order + 1,
        "is_adhoc": True,
    })
    enrollment.journey = journey
    enrollment.updated_at = datetime.utcnow()
    await enrollment.save()
    return enrollment_to_response(enrollment)


@router.delete("/{enrollment_id}/journey/{step_id}")
async def delete_journey_step(
    enrollment_id: str,
    step_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Remove a step from this enrollment's journey."""
    enrollment = await _get_enrollment_or_404(enrollment_id)
    _require_spoc_or_admin(enrollment, current_user)

    journey = enrollment.journey or []
    new_journey = [s for s in journey if s.get("step_id") != step_id]
    if len(new_journey) == len(journey):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Journey step not found")
    enrollment.journey = new_journey
    enrollment.updated_at = datetime.utcnow()
    await enrollment.save()
    return enrollment_to_response(enrollment)


class JourneyStopRequest(BaseModel):
    reason: Optional[str] = None


class DncRequest(BaseModel):
    do_not_contact: bool = True
    reason: Optional[str] = None


class JourneyFlagRequest(BaseModel):
    flag: str = "trimester_contradiction"
    note: Optional[str] = None


class ReclassifyRequest(BaseModel):
    target: str        # "PreConception" | "Antenatal" | "Outreach"


@router.post("/{enrollment_id}/journey/stop")
async def stop_enrollment_journey(
    enrollment_id: str,
    body: JourneyStopRequest,
    current_user: dict = Depends(get_current_user),
):
    """Stop the care journey (SPOC/admin): cancel pending steps, keep history."""
    enrollment = await _get_enrollment_or_404(enrollment_id)
    _require_spoc_or_admin(enrollment, current_user)
    enrollment.journey = stop_journey_steps(enrollment.journey or [])
    enrollment.journey_status = "stopped"
    enrollment.journey_stopped_reason = body.reason
    enrollment.journey_stopped_by = current_user["user_id"]
    enrollment.journey_stopped_by_name = current_user.get("full_name", current_user["email"])
    enrollment.journey_stopped_at = datetime.utcnow()
    enrollment.updated_at = datetime.utcnow()
    await enrollment.save()
    return enrollment_to_response(enrollment)


@router.post("/{enrollment_id}/dnc")
async def set_enrollment_dnc(
    enrollment_id: str,
    body: DncRequest,
    current_user: dict = Depends(get_current_user),
):
    """Set/clear Do-Not-Contact (SPOC/admin). Setting it hard-stops the journey."""
    enrollment = await _get_enrollment_or_404(enrollment_id)
    _require_spoc_or_admin(enrollment, current_user)
    enrollment.do_not_contact = bool(body.do_not_contact)
    if enrollment.do_not_contact:
        enrollment.dnc_reason = body.reason
        enrollment.dnc_set_by = current_user["user_id"]
        enrollment.dnc_at = datetime.utcnow()
        enrollment.journey = stop_journey_steps(enrollment.journey or [])
        enrollment.journey_status = "stopped"
        enrollment.journey_stopped_reason = enrollment.journey_stopped_reason or "Do Not Contact"
        enrollment.journey_stopped_at = datetime.utcnow()
    else:
        enrollment.dnc_reason = None
    enrollment.updated_at = datetime.utcnow()
    await enrollment.save()
    return enrollment_to_response(enrollment)


@router.post("/{enrollment_id}/journey/flag")
async def flag_enrollment_journey(
    enrollment_id: str,
    body: JourneyFlagRequest,
    current_user: dict = Depends(get_current_user),
):
    """Agent flags the record to Admin (e.g. Antenatal + 'Not Conceived')."""
    enrollment = await _get_enrollment_or_404(enrollment_id)
    _require_spoc_or_admin(enrollment, current_user)
    enrollment.journey_flag = body.flag
    enrollment.journey_flag_note = body.note
    enrollment.journey_flagged_at = datetime.utcnow()
    enrollment.updated_at = datetime.utcnow()
    await enrollment.save()
    return enrollment_to_response(enrollment)


@router.post("/{enrollment_id}/journey/reclassify")
async def reclassify_enrollment_journey(
    enrollment_id: str,
    body: ReclassifyRequest,
    current_user: dict = Depends(get_current_admin),
):
    """
    Admin reclassifies a flagged record's journey (spec 4e):
      - PreConception / Antenatal -> set service_enrolled to the target and rebuild
        the care journey (preserving completed steps).
      - Outreach -> stop the care journey; the lead is handled via outreach.
    Clears the agent flag.
    """
    target = (body.target or "").strip()
    enrollment = await _get_enrollment_or_404(enrollment_id)

    if target.lower() in ("preconception", "antenatal", "maternitywellness"):
        canonical = next(
            (s for s in ("PreConception", "Antenatal", "MaternityWellness")
             if s.lower() == target.lower()), target
        )
        enrollment.service_enrolled = canonical
        enrollment.journey_classification = canonical
        enrollment.journey_status = "active"
        enrollment.journey = await reinstantiate_care_journey(enrollment)
    elif target.lower() == "outreach":
        enrollment.journey_classification = "Outreach"
        enrollment.journey = stop_journey_steps(enrollment.journey or [])
        enrollment.journey_status = "stopped"
        enrollment.journey_stopped_reason = "Reclassified to outreach"
        enrollment.journey_stopped_at = datetime.utcnow()
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="target must be PreConception, Antenatal, MaternityWellness or Outreach",
        )

    enrollment.journey_flag = None
    enrollment.journey_flag_note = None
    enrollment.updated_at = datetime.utcnow()
    enrollment.last_modified_by = current_user["user_id"]
    await enrollment.save()
    return enrollment_to_response(enrollment)


@router.post("/{enrollment_id}/convert-to-antenatal")
async def convert_to_antenatal(
    enrollment_id: str,
    current_user: dict = Depends(get_current_user),
):
    """
    PreConception -> Antenatal happy path (SPOC/admin): spawn a new pre-filled
    Antenatal lead (carrying contact details, linked back to this enrollment) and
    stop the PreConception keep-in-touch loop (reason "Conceived").
    """
    enrollment = await _get_enrollment_or_404(enrollment_id)
    _require_spoc_or_admin(enrollment, current_user)

    if enrollment.converted_to_lead_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Already converted to lead {enrollment.converted_to_lead_id}",
        )

    db = get_database()
    new_lead_id = await generate_lead_id(db)
    new_lead = Lead(
        lead_id=new_lead_id,
        lead_creation_date=date.today(),
        status=LeadStatus.ENQUIRY_LEAD.value,
        name=enrollment.name or enrollment.subscriber_name or "Unknown",
        email=enrollment.email,
        phone_number=enrollment.phone_number,
        employee_id=enrollment.employee_id or None,
        uhid=enrollment.uhid,
        service_requested="Antenatal",
        assigned_to=enrollment.assigned_to,
        assigned_to_name=enrollment.assigned_to_name,
        assigned_date=datetime.utcnow(),
        hclhc_spoc=enrollment.hclhc_spoc,
        converted_from_enrollment_id=enrollment.enrollment_id,
        created_by=current_user["user_id"],
    )
    await new_lead.insert()

    # Stop the PreConception keep-in-touch loop, keep history.
    enrollment.journey = stop_journey_steps(enrollment.journey or [])
    enrollment.journey_status = "stopped"
    enrollment.journey_stopped_reason = "Conceived"
    enrollment.journey_stopped_by = current_user["user_id"]
    enrollment.journey_stopped_by_name = current_user.get("full_name", current_user["email"])
    enrollment.journey_stopped_at = datetime.utcnow()
    enrollment.converted_to_lead_id = new_lead_id
    enrollment.updated_at = datetime.utcnow()
    await enrollment.save()

    return {
        "message": f"Created Antenatal lead {new_lead_id} from enrollment {enrollment_id}",
        "lead_id": new_lead_id,
        "enrollment": enrollment_to_response(enrollment),
    }


@router.get("/{enrollment_id}/audit")
async def get_enrollment_audit_trail(
    enrollment_id: str,
    current_user: dict = Depends(get_current_admin)
):
    """Get audit trail for an enrollment (Admin only)"""
    # Verify enrollment exists
    enrollment = await Enrollment.find_one(
        Enrollment.enrollment_id == enrollment_id
    )
    if not enrollment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Enrollment not found"
        )

    # Get audit logs sorted by timestamp descending (most recent first)
    audit_logs = await EnrollmentAuditLog.find(
        EnrollmentAuditLog.enrollment_id == enrollment_id
    ).sort("-timestamp").to_list()

    return {
        "enrollment_id": enrollment_id,
        "audit_trail": [
            {
                "id": str(log.id),
                "user_email": log.user_email,
                "user_name": log.user_name,
                "action": log.action.value,
                "changes": log.changes,
                "timestamp": log.timestamp.isoformat()
            }
            for log in audit_logs
        ]
    }
